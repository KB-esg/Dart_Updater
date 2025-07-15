import os
from datetime import datetime, timedelta
import json
import time
import re
import gspread
from google.oauth2.service_account import Credentials
import OpenDartReader
import requests
from bs4 import BeautifulSoup
try:
    from html_table_parser import parser_functions as parser
except ImportError:
    try:
        import html_table_parser
        # html_table_parser가 있지만 make2d가 없는 경우 직접 구현
        class TableParser:
            @staticmethod
            def make2d(table):
                """BeautifulSoup table을 2D 리스트로 변환"""
                rows = []
                for tr in table.find_all('tr'):
                    row = []
                    for td in tr.find_all(['td', 'th']):
                        # colspan과 rowspan 처리
                        colspan = int(td.get('colspan', 1))
                        text = td.get_text(strip=True)
                        
                        # colspan만큼 반복
                        for _ in range(colspan):
                            row.append(text)
                    if row:
                        rows.append(row)
                return rows
        
        parser = TableParser()
        
    except ImportError:
        # html_table_parser가 완전히 없는 경우
        class TableParser:
            @staticmethod
            def make2d(table):
                """BeautifulSoup table을 2D 리스트로 변환하는 대체 함수"""
                rows = []
                for tr in table.find_all('tr'):
                    row = []
                    for td in tr.find_all(['td', 'th']):
                        row.append(td.get_text(strip=True))
                    if row:
                        rows.append(row)
                return rows
        
        parser = TableParser()
import pandas as pd
from openpyxl import load_workbook
from playwright.sync_api import sync_playwright
import shutil
from tqdm import tqdm

class DartDualUpdater:
    """DART XBRL Excel 다운로드 + HTML 스크래핑 통합 시스템 (안전한 버전)"""
    
    # HTML 스크래핑 대상 시트 (재무제표 관련 제외)
    HTML_TARGET_SHEETS = [
        'I. 회사의 개요', 'II. 사업의 내용', '1. 사업의 개요', '2. 주요 제품 및 서비스',
        '3. 원재료 및 생산설비', '4. 매출 및 수주상황', '5. 위험관리 및 파생거래',
        '6. 주요계약 및 연구활동', '7. 기타 참고 사항', '1. 요약재무정보',
        '6. 배당에 관한 사항', '8. 기타 재무에 관한 사항', 'VII. 주주에 관한 사항',
        'VIII. 임원 및 직원 등에 관한 사항', 'X. 대주주 등과의 거래내용',
        'XI. 그 밖에 투자자 보호를 위하여 필요한 사항'
    ]
    
    def __init__(self, company_config):
        """초기화"""
        self.corp_code = company_config['corp_code']
        self.company_name = company_config['company_name']
        self.spreadsheet_var_name = company_config['spreadsheet_var']
        
        # 환경변수 확인
        self._check_environment_variables()
        
        # Google Sheets 설정 (재시도 로직 포함)
        self.credentials = self._get_google_credentials()
        self.gc = gspread.authorize(self.credentials)
        self.workbook = self._connect_to_spreadsheet_with_retry()
        
        # DART API 설정
        self.dart = OpenDartReader(os.environ['DART_API_KEY'])
        
        # 텔레그램 설정
        self.telegram_bot_token = os.environ.get('TELEGRAM_BOT_TOKEN')
        self.telegram_channel_id = os.environ.get('TELEGRAM_CHANNEL_ID')
        
        # 다운로드 폴더 설정 (XBRL용)
        self.download_dir = os.path.join(os.getcwd(), 'downloads')
        os.makedirs(self.download_dir, exist_ok=True)
        
        # 처리 결과 추적
        self.results = {
            'total_reports': 0,
            'xbrl': {
                'downloaded_files': [],
                'uploaded_sheets': [],
                'failed_downloads': [],
                'failed_uploads': [],
                'excel_files': {}
            },
            'html': {
                'processed_sheets': [],
                'failed_sheets': []
            }
        }
        
        # 현재 처리 중인 보고서 정보
        self.current_report = None
        
        # Archive 시트 행 영역 매핑 설정
        self._setup_archive_row_mapping()

    def _check_environment_variables(self):
        """환경변수 확인"""
        print("🔍 환경변수 확인:")
        required_vars = ['DART_API_KEY', 'GOOGLE_CREDENTIALS', self.spreadsheet_var_name]
        
        for var in required_vars:
            if var in os.environ:
                value = os.environ[var]
                masked_value = f"{value[:6]}...{value[-4:]}" if len(value) > 20 else f"{value[:-2]}**"
                print(f"✅ {var}: {masked_value} (길이: {len(value)})")
            else:
                raise ValueError(f"❌ {var} 환경변수가 설정되지 않았습니다.")

    def _get_google_credentials(self):
        """Google Sheets 인증 설정"""
        creds_json = json.loads(os.environ['GOOGLE_CREDENTIALS'])
        scopes = [
            'https://www.googleapis.com/auth/spreadsheets',
            'https://www.googleapis.com/auth/drive'
        ]
        return Credentials.from_service_account_info(creds_json, scopes=scopes)

    def _connect_to_spreadsheet_with_retry(self, max_retries=5):
        """Google Spreadsheet 연결 (재시도 로직 포함)"""
        for attempt in range(max_retries):
            try:
                print(f"📊 Google Spreadsheet 연결 시도 {attempt + 1}/{max_retries}...")
                workbook = self.gc.open_by_key(os.environ[self.spreadsheet_var_name])
                print(f"✅ Google Spreadsheet 연결 성공!")
                return workbook
                
            except gspread.exceptions.APIError as e:
                error_code = str(e).split('[')[1].split(']')[0] if '[' in str(e) and ']' in str(e) else 'unknown'
                
                if error_code in ['503', '502', '500', '429']:
                    wait_time = min(30 * (2 ** attempt), 300)  # 지수 백오프, 최대 5분
                    print(f"⚠️ Google Sheets API 오류 {error_code}: {str(e)}")
                    
                    if attempt < max_retries - 1:
                        print(f"⏳ {wait_time}초 후 재시도... (시도 {attempt + 1}/{max_retries})")
                        time.sleep(wait_time)
                        continue
                    else:
                        print(f"❌ 최종 실패: Google Sheets 연결 불가")
                        raise e
                else:
                    # 인증 오류나 권한 오류 등은 재시도하지 않음
                    print(f"❌ Google Sheets 연결 실패 (재시도 불가): {str(e)}")
                    raise e
                    
            except Exception as e:
                print(f"⚠️ 예상치 못한 오류 (시도 {attempt + 1}/{max_retries}): {str(e)}")
                if attempt < max_retries - 1:
                    wait_time = min(15 * (attempt + 1), 60)
                    print(f"⏳ {wait_time}초 후 재시도...")
                    time.sleep(wait_time)
                    continue
                else:
                    print(f"❌ 최종 실패: Google Sheets 연결 불가")
                    raise e
        
        raise Exception("Google Spreadsheet 연결에 실패했습니다.")

    def _execute_sheets_operation_with_retry(self, operation, *args, max_retries=3, **kwargs):
        """Google Sheets 작업 실행 (재시도 로직 포함)"""
        for attempt in range(max_retries):
            try:
                return operation(*args, **kwargs)
                
            except gspread.exceptions.APIError as e:
                error_code = str(e).split('[')[1].split(']')[0] if '[' in str(e) and ']' in str(e) else 'unknown'
                
                if error_code in ['503', '502', '500', '429']:
                    if attempt < max_retries - 1:
                        wait_time = min(30 * (2 ** attempt), 120)  # 지수 백오프, 최대 2분
                        print(f"⚠️ Google Sheets API 오류 {error_code} (시도 {attempt + 1}/{max_retries})")
                        print(f"⏳ {wait_time}초 후 재시도...")
                        time.sleep(wait_time)
                        continue
                    else:
                        print(f"❌ Google Sheets 작업 최종 실패: {str(e)}")
                        raise e
                else:
                    print(f"❌ Google Sheets 작업 실패 (재시도 불가): {str(e)}")
                    raise e
                    
            except Exception as e:
                if attempt < max_retries - 1:
                    wait_time = 10 * (attempt + 1)
                    print(f"⚠️ 예상치 못한 오류 (시도 {attempt + 1}/{max_retries}): {str(e)}")
                    print(f"⏳ {wait_time}초 후 재시도...")
                    time.sleep(wait_time)
                    continue
                else:
                    print(f"❌ 작업 최종 실패: {str(e)}")
                    raise e
        
        raise Exception("Google Sheets 작업 실행에 실패했습니다.")

    def _setup_archive_row_mapping(self):
        """Archive 시트의 행 영역 매핑 설정"""
        # 재무제표 Archive 시트 행 매핑
        self.financial_row_mapping = {
            'consolidated': {
                'D210000': {'start': 7, 'end': 80, 'name': '연결 재무상태표'},
                'D431410': {'start': 81, 'end': 140, 'name': '연결 손익계산서'},
                'D520000': {'start': 141, 'end': 200, 'name': '연결 현금흐름표'},
                'D610000': {'start': 201, 'end': 250, 'name': '연결 자본변동표'}
            },
            'standalone': {
                'D210005': {'start': 257, 'end': 330, 'name': '별도 재무상태표'},
                'D431415': {'start': 331, 'end': 390, 'name': '별도 손익계산서'},
                'D520005': {'start': 391, 'end': 450, 'name': '별도 현금흐름표'},
                'D610005': {'start': 451, 'end': 500, 'name': '별도 자본변동표'}
            }
        }

    def run(self):
        """메인 실행 함수 (문서별 순차 처리)"""
        print(f"\n🚀 {self.company_name}({self.corp_code}) DART 통합 업데이트 시작")
        print("📊 업데이트 모드: 문서별 XBRL → Archive → HTML → Archive 순서")
        
        # 단위 정보 출력
        number_unit = os.environ.get('NUMBER_UNIT', 'million')
        unit_text = {
            'million': '백만원',
            'hundred_million': '억원',
            'billion': '십억원'
        }.get(number_unit, '백만원')
        print(f"💰 숫자 표시 단위: {unit_text}")
        
        # 1. 보고서 목록 조회
        reports = self._get_recent_reports()
        if reports.empty:
            print("📭 최근 보고서가 없습니다.")
            return
        
        print(f"📋 발견된 보고서: {len(reports)}개")
        self.results['total_reports'] = len(reports)
        
        # 2. 문서별로 순차 처리 (XBRL → XBRL Archive → HTML → HTML Archive)
        with sync_playwright() as p:
            browser = p.chromium.launch(
                headless=True,
                args=[
                    '--disable-blink-features=AutomationControlled',
                    '--no-sandbox',
                    '--disable-setuid-sandbox',
                    '--disable-dev-shm-usage'
                ]
            )
            context = browser.new_context(
                accept_downloads=True,
                locale='ko-KR',
                viewport={'width': 1920, 'height': 1080},
                user_agent='Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
            )
            
            try:
                with tqdm(total=len(reports), desc="문서별 처리", unit="건") as pbar:
                    for _, report in reports.iterrows():
                        print(f"\n{'='*60}")
                        print(f"📄 문서 처리 시작: {report['report_nm']} (접수번호: {report['rcept_no']})")
                        print(f"{'='*60}")
                        
                        # Step 1: XBRL Excel 다운로드
                        print("\n🔸 Step 1: XBRL Excel 다운로드")
                        self._process_xbrl_report(context, report)
                        
                        # Step 2: XBRL Archive 업데이트 (방금 다운로드한 파일)
                        if self.results['xbrl']['excel_files']:
                            print("\n🔸 Step 2: XBRL Archive 업데이트")
                            if os.environ.get('ENABLE_ARCHIVE_UPDATE', 'true').lower() == 'true':
                                self._update_xbrl_archive_for_current_report()
                        
                        # Step 3: HTML 스크래핑 (단순한 방식으로 변경)
                        print("\n🔸 Step 3: HTML 스크래핑")
                        self._process_html_report_simple(report['rcept_no'])
                        
                        # Step 4: HTML Archive 업데이트
                        print("\n🔸 Step 4: HTML Archive 업데이트")
                        if os.environ.get('ENABLE_HTML_ARCHIVE', 'true').lower() == 'true':
                            self._update_html_archive_for_current_report()
                        
                        # 파일 정리 (다음 문서 처리 전)
                        self._cleanup_current_downloads()
                        
                        print(f"✅ 문서 처리 완료: {report['rcept_no']}")
                        pbar.update(1)
                        
                        # 문서 간 대기 (API 제한 회피)
                        time.sleep(3)
                    
            finally:
                browser.close()
        
        # 5. 결과 요약
        self._print_summary()
        
        # 6. 최종 정리
        self._cleanup_downloads()

    def _get_recent_reports(self):
        """최근 보고서 목록 조회"""
        start_date, end_date = self._get_date_range()
        return self.dart.list(self.corp_code, start_date, end_date, kind='A', final='T')

    def _get_date_range(self):
        """날짜 범위 계산"""
        manual_start = os.environ.get('MANUAL_START_DATE')
        manual_end = os.environ.get('MANUAL_END_DATE')
        
        if manual_start and manual_end:
            print(f"📅 수동 설정 날짜: {manual_start} ~ {manual_end}")
            return manual_start, manual_end
        
        end_date = datetime.now()
        start_date = end_date - timedelta(days=90)
        date_range = start_date.strftime('%Y%m%d'), end_date.strftime('%Y%m%d')
        print(f"📅 기본 날짜 범위 (최근 3개월): {date_range[0]} ~ {date_range[1]}")
        return date_range

    # === XBRL 관련 메서드 ===
    def _process_xbrl_report(self, context, report):
        """XBRL 보고서 처리"""
        print(f"\n📄 XBRL 처리: {report['report_nm']} (접수번호: {report['rcept_no']})")
        
        self.current_report = report
        page = context.new_page()
        
        try:
            viewer_url = f"https://opendart.fss.or.kr/xbrl/viewer/main.do?rcpNo={report['rcept_no']}"
            print(f"🌐 페이지 열기: {viewer_url}")
            
            page.goto(viewer_url, wait_until='networkidle', timeout=60000)
            page.wait_for_timeout(2000)
            
            download_button = page.locator('button.btnDown').first
            if not download_button.is_visible():
                print("⚠️ 다운로드 버튼을 찾을 수 없습니다.")
                self.results['xbrl']['failed_downloads'].append(report['rcept_no'])
                return
            
            print("🖱️ 다운로드 버튼 클릭")
            
            with page.expect_popup() as popup_info:
                download_button.click()
            
            popup = popup_info.value
            popup.wait_for_load_state('networkidle')
            
            self._download_excel_files(popup, report['rcept_no'])
            popup.close()
            
        except Exception as e:
            print(f"❌ XBRL 처리 실패: {str(e)}")
            self.results['xbrl']['failed_downloads'].append(report['rcept_no'])
        finally:
            page.close()

    def _download_excel_files(self, popup_page, rcept_no):
        """팝업 페이지에서 Excel 파일 다운로드"""
        try:
            popup_page.wait_for_timeout(2000)
            print(f"📍 팝업 페이지 URL: {popup_page.url}")
            
            download_links = popup_page.locator('a.btnFile')
            link_count = download_links.count()
            print(f"📄 다운로드 가능한 파일 수: {link_count}개")
            
            # 재무제표 다운로드
            if link_count >= 1:
                print("📥 재무제표 다운로드 중...")
                
                with popup_page.expect_download() as download_info:
                    download_links.nth(0).click()
                
                download = download_info.value
                file_path = os.path.join(self.download_dir, f"재무제표_{rcept_no}.xlsx")
                download.save_as(file_path)
                
                print(f"✅ 재무제표 다운로드 완료: {file_path}")
                self.results['xbrl']['downloaded_files'].append(file_path)
                self.results['xbrl']['excel_files']['financial'] = file_path
                
                self._upload_excel_to_sheets(file_path, "재무제표", rcept_no)
                popup_page.wait_for_timeout(2000)
            
            # 재무제표주석 다운로드
            if link_count >= 2:
                print("📥 재무제표주석 다운로드 중...")
                
                with popup_page.expect_download() as download_info:
                    download_links.nth(1).click()
                
                download = download_info.value
                file_path = os.path.join(self.download_dir, f"재무제표주석_{rcept_no}.xlsx")
                download.save_as(file_path)
                
                print(f"✅ 재무제표주석 다운로드 완료: {file_path}")
                self.results['xbrl']['downloaded_files'].append(file_path)
                self.results['xbrl']['excel_files']['notes'] = file_path
                
                self._upload_excel_to_sheets(file_path, "재무제표주석", rcept_no)
                
        except Exception as e:
            print(f"❌ Excel 다운로드 실패: {str(e)}")
            self.results['xbrl']['failed_downloads'].append(f"Excel_{rcept_no}")

    # === HTML 스크래핑 관련 메서드 (완전히 재작성) ===
    def _process_html_report_simple(self, rcept_no):
        """HTML 보고서 처리 (단순화된 버전)"""
        try:
            print(f"\n🌐 HTML 처리: 보고서 접수번호 {rcept_no}")
            
            # 보고서 하위 문서 목록 조회
            report_index = self._get_report_index_with_retry(rcept_no)
            if report_index is None or report_index.empty:
                print("⚠️ 보고서 하위 문서를 찾을 수 없습니다.")
                return
            
            # HTML 대상 시트만 필터링
            target_docs = report_index[report_index['title'].isin(self.HTML_TARGET_SHEETS)]
            
            print(f"📝 처리할 HTML 문서: {len(target_docs)}개")
            
            for _, doc in target_docs.iterrows():
                self._update_worksheet_simple(doc['title'], doc['url'])
                time.sleep(2)  # 각 문서 간 대기
                
        except Exception as e:
            print(f"❌ HTML 보고서 처리 실패: {str(e)}")

    def _get_report_index_with_retry(self, rcept_no, max_retries=3):
        """보고서 인덱스 조회 (재시도 포함)"""
        for attempt in range(max_retries):
            try:
                report_index = self.dart.sub_docs(rcept_no)
                if report_index is not None and not report_index.empty:
                    return report_index
                else:
                    print(f"⚠️ 시도 {attempt + 1}: 보고서 인덱스가 비어있음")
            except Exception as e:
                print(f"⚠️ 시도 {attempt + 1}: 보고서 인덱스 조회 실패 - {str(e)}")
                
            if attempt < max_retries - 1:
                wait_time = (attempt + 1) * 5
                print(f"⏳ {wait_time}초 후 재시도...")
                time.sleep(wait_time)
        
        return None

    def _update_worksheet_simple(self, sheet_name, url):
        """워크시트 업데이트 (기존 삼성SDS 코드 방식 적용)"""
        max_retries = 3
        retry_delay = 5
        
        for attempt in range(max_retries):
            try:
                print(f"📄 처리 중: {sheet_name} (시도 {attempt + 1}/{max_retries})")
                
                # 워크시트 가져오기 또는 생성
                try:
                    worksheet = self._execute_sheets_operation_with_retry(
                        self.workbook.worksheet, sheet_name
                    )
                except gspread.exceptions.WorksheetNotFound:
                    worksheet = self._execute_sheets_operation_with_retry(
                        self.workbook.add_worksheet, sheet_name, 1000, 10
                    )
                    print(f"🆕 새 시트 생성: {sheet_name}")
                    time.sleep(2)
                
                # HTTP 요청 (기존 방식 사용)
                response = requests.get(url, timeout=30)
                
                if response.status_code == 200:
                    content_length = len(response.content)
                    print(f"📥 콘텐츠 크기: {content_length:,} bytes")
                    
                    if content_length < 100:
                        print(f"⚠️ 콘텐츠가 너무 작습니다: {content_length} bytes")
                        if attempt < max_retries - 1:
                            time.sleep(retry_delay)
                            continue
                    
                    # HTML 처리 (완전히 단순화)
                    success = self._process_html_content_simple(worksheet, response.text)
                    
                    if success:
                        print(f"✅ HTML 시트 업데이트 완료: {sheet_name}")
                        self.results['html']['processed_sheets'].append(sheet_name)
                        return
                    else:
                        if attempt < max_retries - 1:
                            print(f"⏳ {retry_delay}초 후 재시도...")
                            time.sleep(retry_delay)
                            continue
                        else:
                            print(f"❌ 최종 실패: {sheet_name}")
                            self.results['html']['failed_sheets'].append(sheet_name)
                            return
                else:
                    print(f"⚠️ HTTP {response.status_code}: {sheet_name}")
                    
            except requests.exceptions.Timeout:
                print(f"⚠️ 타임아웃 (시도 {attempt + 1}/{max_retries}): {sheet_name}")
            except requests.exceptions.ConnectionError as e:
                print(f"⚠️ 연결 오류 (시도 {attempt + 1}/{max_retries}): {sheet_name} - {str(e)}")
            except Exception as e:
                print(f"❌ HTML 워크시트 업데이트 실패 ({sheet_name}): {str(e)}")
                if attempt == max_retries - 1:
                    self.results['html']['failed_sheets'].append(sheet_name)
                return
            
            if attempt < max_retries - 1:
                print(f"⏳ {retry_delay:.1f}초 후 재시도...")
                time.sleep(retry_delay)
                retry_delay *= 1.5
        
        print(f"❌ 최종 실패: {sheet_name}")
        self.results['html']['failed_sheets'].append(sheet_name)

    def _process_html_content_simple(self, worksheet, html_content):
        """HTML 내용 처리 (기존 삼성SDS 방식 적용)"""
        try:
            # BeautifulSoup으로 파싱
            soup = BeautifulSoup(html_content, 'html.parser')
            tables = soup.find_all("table")
            
            # 워크시트 클리어
            self._execute_sheets_operation_with_retry(worksheet.clear)
            
            # 메타데이터 추가
            meta_data = [
                [f"업데이트: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"],
                [f"보고서: {self.current_report.get('rcept_no', '') if self.current_report else ''}"],
                [f"회사: {self.company_name}"],
                []
            ]
            
            # HTML 컨텐츠 검증
            if not html_content or len(html_content.strip()) < 100:
                print("⚠️ HTML 컨텐츠가 비어있거나 너무 짧습니다.")
                error_data = meta_data + [["오류: HTML 컨텐츠가 비어있음"]]
                return self._simple_batch_update(worksheet, error_data)
            
            # 테이블 데이터 수집
            all_data = meta_data.copy()
            
            if tables:
                for table_idx, table in enumerate(tables):
                    try:
                        table_data = parser.make2d(table)
                        if table_data and len(table_data) > 0:
                            # 테이블 헤더 추가
                            all_data.append([f"=== 테이블 {table_idx + 1} ==="])
                            all_data.extend(table_data)
                            all_data.append([])  # 구분을 위한 빈 행
                    except Exception as e:
                        print(f"⚠️ 테이블 {table_idx + 1} 파싱 오류: {str(e)}")
                        all_data.append([f"테이블 {table_idx + 1} 파싱 오류"])
                        continue
            else:
                # 테이블이 없으면 텍스트 내용 추출
                text_content = soup.get_text()
                if text_content and len(text_content.strip()) > 50:
                    lines = text_content.split('\n')
                    for line in lines[:50]:  # 최대 50줄
                        clean_line = line.strip()
                        if clean_line and len(clean_line) > 2:
                            all_data.append([clean_line[:200]])  # 최대 200자
                else:
                    all_data.append(["텍스트 컨텐츠를 찾을 수 없습니다."])
            
            # 단순한 배치 업데이트 (기존 SDS 방식)
            return self._simple_batch_update(worksheet, all_data)
            
        except Exception as e:
            print(f"❌ HTML 콘텐츠 처리 실패: {str(e)}")
            try:
                # 최소한의 오류 정보 저장
                error_data = [
                    [f"오류 발생: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"],
                    [f"오류 내용: {str(e)[:100]}"],
                    [f"보고서: {self.current_report.get('rcept_no', '') if self.current_report else ''}"]
                ]
                self._simple_batch_update(worksheet, error_data)
            except:
                pass
            return False

    def _simple_batch_update(self, worksheet, all_data):
        """단순한 배치 업데이트 (기존 SDS 방식)"""
        try:
            if not all_data:
                return False
                
            # 기존 방식: append_rows 사용
            BATCH_SIZE = 100
            for i in range(0, len(all_data), BATCH_SIZE):
                batch = all_data[i:i + BATCH_SIZE]
                try:
                    # 각 행의 길이를 동일하게 맞춤
                    max_cols = max(len(row) if row else 1 for row in batch)
                    normalized_batch = []
                    for row in batch:
                        if not row:
                            normalized_row = [''] * max_cols
                        else:
                            normalized_row = row + [''] * (max_cols - len(row))
                        normalized_batch.append(normalized_row)
                    
                    # append_rows 사용 (기존 방식)
                    worksheet.append_rows(normalized_batch)
                    print(f"배치 업데이트 완료: {i+1}~{min(i+BATCH_SIZE, len(all_data))} 행")
                    time.sleep(1)  # API 제한 회피
                    
                except gspread.exceptions.APIError as e:
                    if 'Quota exceeded' in str(e):
                        print("할당량 제한 도달. 60초 대기 후 재시도...")
                        time.sleep(60)
                        worksheet.append_rows(normalized_batch)
                    else:
                        print(f"⚠️ 배치 {i+1}-{i+len(batch)} 업데이트 실패: {str(e)}")
                        continue
                except Exception as e:
                    print(f"⚠️ 배치 {i+1}-{i+len(batch)} 업데이트 실패: {str(e)}")
                    continue
                    
            return True
            
        except Exception as e:
            print(f"❌ 배치 업데이트 전체 실패: {str(e)}")
            return False

    def _update_html_archive_for_current_report(self):
        """현재 보고서의 HTML Archive 업데이트 (최적화된 DataFrame 방식)"""
        print("📊 현재 문서 HTML Archive 업데이트 중...")
        
        try:
            # Dart_Archive 시트 접근
            try:
                archive = self._execute_sheets_operation_with_retry(
                    self.workbook.worksheet, 'Dart_Archive'
                )
            except gspread.exceptions.WorksheetNotFound:
                print("⚠️ Dart_Archive 시트를 찾을 수 없습니다.")
                return
                
            sheet_values = self._execute_sheets_operation_with_retry(
                archive.get_all_values
            )
            
            if not sheet_values:
                print("⚠️ Dart_Archive 시트가 비어있습니다")
                return
            
            last_col = len(sheet_values[0]) if sheet_values[0] else 0
            
            # 마지막 열에 데이터가 있는지 확인
            try:
                control_value = self._execute_sheets_operation_with_retry(
                    archive.cell, 1, last_col
                ).value if last_col > 0 else None
                
                if control_value:
                    last_col += 1
            except:
                last_col += 1
            
            # 최적화된 방식으로 처리
            self._process_archive_data_optimized(archive, 10, last_col)
            print("✅ 현재 문서 HTML Archive 업데이트 완료")
            
        except Exception as e:
            print(f"❌ 현재 문서 HTML Archive 업데이트 실패: {str(e)}")

    def _process_archive_data_optimized(self, archive, start_row, last_col):
        """최적화된 아카이브 데이터 처리 (DataFrame 기반 + 단일 배치 업데이트)"""
        try:
            current_cols = archive.col_count
            current_col_letter = self._get_column_letter(current_cols)
            target_col_letter = self._get_column_letter(last_col)
            
            print(f"시작 행: {start_row}, 대상 열: {last_col} ({target_col_letter})")
            print(f"현재 시트 열 수: {current_cols} ({current_col_letter})")
            
            # 필요한 경우 시트 크기 조정
            if last_col >= current_cols:
                new_cols = last_col + 5
                try:
                    print(f"시트 크기를 {current_cols}({current_col_letter})에서 {new_cols}({self._get_column_letter(new_cols)})로 조정합니다.")
                    self._execute_sheets_operation_with_retry(
                        archive.resize, rows=archive.row_count, cols=new_cols
                    )
                    time.sleep(2)
                    print("시트 크기 조정 완료")
                except Exception as e:
                    print(f"시트 크기 조정 중 오류 발생: {str(e)}")
                    raise

            # 1단계: Archive에서 검색 작업 목록 읽기 (1번 API 호출)
            print("📋 Archive에서 검색 작업 목록 로드 중...")
            all_rows = self._execute_sheets_operation_with_retry(archive.get_all_values)
            
            # 검색 작업을 시트별로 그룹화
            search_tasks_by_sheet = {}
            for row_idx in range(start_row - 1, len(all_rows)):
                if len(all_rows[row_idx]) < 5:
                    continue
                    
                sheet_name = all_rows[row_idx][0]
                keyword = all_rows[row_idx][1]
                n = all_rows[row_idx][2]
                x = all_rows[row_idx][3]
                y = all_rows[row_idx][4]
                
                if not sheet_name or not keyword:
                    continue
                
                try:
                    search_task = {
                        'archive_row': row_idx + 1,
                        'keyword': keyword,
                        'n': int(n),
                        'x': int(x),
                        'y': int(y)
                    }
                    
                    if sheet_name not in search_tasks_by_sheet:
                        search_tasks_by_sheet[sheet_name] = []
                    search_tasks_by_sheet[sheet_name].append(search_task)
                    
                except (ValueError, TypeError):
                    print(f"⚠️ 행 {row_idx + 1}: 잘못된 검색 파라미터")
                    continue
            
            print(f"📊 총 {len(search_tasks_by_sheet)}개 시트에서 검색 작업 수행")
            
            # 2단계: 시트별로 DataFrame 로드 및 모든 검색 수행
            all_results = {}  # {archive_row: value}
            
            for sheet_name, tasks in search_tasks_by_sheet.items():
                print(f"\n🔍 시트 '{sheet_name}' 처리 중 ({len(tasks)}개 키워드)...")
                
                try:
                    # 시트 데이터를 한 번만 로드 (1번 API 호출)
                    search_sheet = self._execute_sheets_operation_with_retry(
                        self.workbook.worksheet, sheet_name
                    )
                    sheet_data = self._execute_sheets_operation_with_retry(
                        search_sheet.get_all_values
                    )
                    
                    if not sheet_data:
                        print(f"⚠️ 시트 '{sheet_name}'가 비어있습니다.")
                        continue
                    
                    # DataFrame으로 변환 (메모리에서 빠른 검색을 위해)
                    df = pd.DataFrame(sheet_data)
                    print(f"📊 시트 크기: {df.shape}")
                    
                    # 해당 시트의 모든 키워드를 메모리에서 검색
                    for task in tasks:
                        try:
                            value = self._search_keyword_in_dataframe(df, task)
                            if value is not None:
                                all_results[task['archive_row']] = value
                                print(f"  ✅ 키워드 '{task['keyword']}' → 값: {str(value)[:50]}")
                            else:
                                print(f"  ❌ 키워드 '{task['keyword']}' 찾을 수 없음")
                        except Exception as e:
                            print(f"  ⚠️ 키워드 '{task['keyword']}' 검색 중 오류: {str(e)}")
                    
                except gspread.exceptions.WorksheetNotFound:
                    print(f"⚠️ 시트 '{sheet_name}'를 찾을 수 없습니다.")
                    continue
                except Exception as e:
                    print(f"⚠️ 시트 '{sheet_name}' 처리 중 오류: {str(e)}")
                    continue
            
            print(f"\n📊 총 {len(all_results)}개 값 발견")
            
            # 3단계: 모든 결과를 한 번에 업데이트 (1번 API 호출)
            if all_results:
                self._execute_single_batch_update(archive, all_results, target_col_letter, last_col)
            else:
                print("⚠️ 업데이트할 데이터가 없습니다.")
            
        except Exception as e:
            error_msg = f"최적화된 아카이브 처리 중 오류 발생: {str(e)}"
            print(error_msg)
            self._send_telegram_message(f"❌ {error_msg}")
            raise e
    
    def _search_keyword_in_dataframe(self, df, task):
        """DataFrame에서 키워드 검색 및 값 추출"""
        try:
            keyword = task['keyword']
            n = task['n']
            x = task['x']
            y = task['y']
            
            # DataFrame에서 키워드 위치 찾기
            keyword_positions = []
            
            # 효율적인 검색: numpy 기반
            mask = (df == keyword)
            positions = mask.stack()
            keyword_positions = [(idx[0], idx[1]) for idx, value in positions.items() if value]
            
            if len(keyword_positions) < n:
                return None
            
            # n번째 키워드 위치
            target_pos = keyword_positions[n - 1]
            target_row = target_pos[0] + y
            target_col = target_pos[1] + x
            
            # 범위 확인
            if (0 <= target_row < df.shape[0] and 0 <= target_col < df.shape[1]):
                value = df.iat[target_row, target_col]
                return self._remove_parentheses(str(value)) if value else ''
            
            return None
            
        except Exception as e:
            print(f"    ⚠️ 키워드 검색 오류: {str(e)}")
            return None
    
    def _execute_single_batch_update(self, archive, results, target_col_letter, last_col):
        """단일 배치로 모든 결과 업데이트"""
        try:
            print(f"📤 단일 배치 업데이트 시작 ({len(results)}개 값)...")
            
            # 결과를 행 번호 순으로 정렬
            sorted_results = sorted(results.items(), key=lambda x: x[0])
            
            min_row = sorted_results[0][0]
            max_row = sorted_results[-1][0]
            
            # 전체 범위의 데이터 배열 생성
            column_data = []
            for row_num in range(min_row, max_row + 1):
                if row_num in results:
                    column_data.append([str(results[row_num])])
                else:
                    column_data.append([''])  # 빈 값
            
            # 단일 배치 업데이트 (1번 API 호출)
            range_label = f'{target_col_letter}{min_row}:{target_col_letter}{max_row}'
            print(f"📋 업데이트 범위: {range_label}")
            
            self._execute_sheets_operation_with_retry(
                archive.update,
                values=column_data,
                range_name=range_label
            )
            
            print(f"✅ 데이터 업데이트 완료: {len(results)}개 값")
            
            # 메타데이터 업데이트 (1번 API 호출)
            today = datetime.now()
            quarter_info = self._get_quarter_info_safe()
            
            meta_updates = [
                {'range': 'J1', 'values': [[today.strftime('%Y-%m-%d')]]},
                {'range': f'{target_col_letter}1', 'values': [['1']]},
                {'range': f'{target_col_letter}5', 'values': [[today.strftime('%Y-%m-%d')]]},
                {'range': f'{target_col_letter}6', 'values': [[quarter_info]]}
            ]
            
            self._execute_sheets_operation_with_retry(
                archive.batch_update, meta_updates
            )
            
            print(f"✅ 메타데이터 업데이트 완료")
            
            # 성공 알림
            message = (
                f"🚀 DART Archive 최적화 업데이트 완료\n\n"
                f"• 종목: {self.company_name} ({self.corp_code})\n"
                f"• 분기: {quarter_info}\n"
                f"• 업데이트 일시: {today.strftime('%Y-%m-%d %H:%M:%S')}\n"
                f"• 처리된 키워드: {len(results)}개\n"
                f"• API 호출 최적화: 시트별 1회 읽기 + 1회 쓰기\n"
                f"• 시트 열: {target_col_letter} (#{last_col})"
            )
            self._send_telegram_message(message)
            
        except Exception as e:
            error_msg = f"단일 배치 업데이트 중 오류: {str(e)}"
            print(error_msg)
            self._send_telegram_message(f"❌ {error_msg}")
            raise e

    def _execute_batch_archive_update_sds_style(self, archive, update_data, target_col_letter, last_col):
        """API 제한을 고려한 최적화된 배치 아카이브 업데이트"""
        try:
            if not update_data:
                return
                
            print(f"📊 총 업데이트할 데이터: {len(update_data)}개")
            
            # 연속된 범위별로 그룹화하여 API 호출 최소화
            sorted_data = sorted(update_data, key=lambda x: x[0])
            batch_groups = []
            current_group = []
            
            for i, (row, value) in enumerate(sorted_data):
                if not current_group:
                    current_group = [(row, value)]
                elif row == sorted_data[i-1][0] + 1:  # 연속된 행
                    current_group.append((row, value))
                else:  # 연속되지 않은 행 - 새 그룹 시작
                    batch_groups.append(current_group)
                    current_group = [(row, value)]
            
            if current_group:
                batch_groups.append(current_group)
            
            print(f"🔄 {len(batch_groups)}개 배치 그룹으로 최적화")
            
            # 각 배치 그룹별로 처리 (API 호출 최소화)
            total_updates = 0
            for group_idx, group in enumerate(batch_groups):
                try:
                    if len(group) == 1:
                        # 단일 셀 업데이트
                        row, value = group[0]
                        range_label = f'{target_col_letter}{row}'
                        values = [[str(value) if value else '']]
                    else:
                        # 연속된 범위 업데이트
                        start_row = group[0][0]
                        end_row = group[-1][0]
                        range_label = f'{target_col_letter}{start_row}:{target_col_letter}{end_row}'
                        values = [[str(item[1]) if item[1] else ''] for item in group]
                    
                    print(f"  📤 배치 {group_idx + 1}/{len(batch_groups)}: {range_label} ({len(group)}개 셀)")
                    
                    self._execute_sheets_operation_with_retry(
                        archive.update, 
                        values=values, 
                        range_name=range_label
                    )
                    
                    total_updates += len(group)
                    
                    # API 제한 회피를 위한 대기
                    if group_idx < len(batch_groups) - 1:  # 마지막 그룹이 아니면
                        if len(batch_groups) > 10:  # 배치가 많으면 더 긴 대기
                            time.sleep(2)
                        else:
                            time.sleep(1)
                
                except Exception as e:
                    print(f"⚠️ 배치 {group_idx + 1} 업데이트 실패: {str(e)}")
                    continue
            
            print(f"✅ 데이터 업데이트 완료: {total_updates}개 셀")
            
            # 메타데이터 업데이트 (단일 배치로 처리)
            today = datetime.now()
            quarter_info = self._get_quarter_info_safe()
            
            meta_updates = [
                {'range': 'J1', 'values': [[today.strftime('%Y-%m-%d')]]},
                {'range': f'{target_col_letter}1', 'values': [['1']]},
                {'range': f'{target_col_letter}5', 'values': [[today.strftime('%Y-%m-%d')]]},
                {'range': f'{target_col_letter}6', 'values': [[quarter_info]]}
            ]
            
            print("📋 메타데이터 업데이트 중...")
            self._execute_sheets_operation_with_retry(
                archive.batch_update, meta_updates
            )
            print(f"✅ 메타데이터 업데이트 완료 (분기: {quarter_info})")
            
            # 텔레그램 알림
            message = (
                f"🔄 DART Archive 업데이트 완료\n\n"
                f"• 종목: {self.company_name} ({self.corp_code})\n"
                f"• 분기: {quarter_info}\n"
                f"• 업데이트 일시: {today.strftime('%Y-%m-%d %H:%M:%S')}\n"
                f"• 처리된 키워드: {total_updates}개\n"
                f"• 배치 그룹: {len(batch_groups)}개\n"
                f"• 시트 열: {target_col_letter} (#{last_col})"
            )
            self._send_telegram_message(message)
            
        except Exception as e:
            error_msg = f"배치 업데이트 중 오류 발생: {str(e)}"
            print(error_msg)
            self._send_telegram_message(f"❌ {error_msg}")
            raise e
            
    def _get_quarter_info_safe(self):
        """안전한 분기 정보 반환 (기존 SDS 방식)"""
        try:
            today = datetime.now()
            three_months_ago = today - timedelta(days=90)
            year = str(three_months_ago.year)[2:]
            quarter = (three_months_ago.month - 1) // 3 + 1
            quarter_text = f"{quarter}Q{year}"
            return quarter_text
        except Exception as e:
            print(f"⚠️ 분기 정보 계산 중 오류: {str(e)}")
            # 기본값 반환
            now = datetime.now()
            quarter = (now.month - 1) // 3 + 1
            year = str(now.year)[2:]
            return f"{quarter}Q{year}"

    def _remove_parentheses(self, value):
        """괄호 내용 제거"""
        if not value or value in ['None', 'nan']:
            return ''
        try:
            return re.sub(r'\s*\(.*?\)\s*', '', str(value)).replace('%', '').strip()
        except:
            return str(value)

    # === 나머지 메서드들 (XBRL 관련) ===
    
    def _upload_excel_to_sheets(self, file_path, file_type, rcept_no):
        """Excel 파일을 Google Sheets에 업로드"""
        try:
            wb = load_workbook(file_path, data_only=True)
            print(f"📊 Excel 파일 열기 완료. 시트 목록: {wb.sheetnames}")
            
            all_sheets_data = {}
            
            print(f"📥 {file_type} 데이터 수집 중...")
            with tqdm(total=len(wb.sheetnames), desc="데이터 수집", unit="시트", leave=False) as pbar:
                for sheet_name in wb.sheetnames:
                    data = []
                    worksheet = wb[sheet_name]
                    for row in worksheet.iter_rows(values_only=True):
                        row_data = [str(cell) if cell is not None else '' for cell in row]
                        if any(row_data):
                            data.append(row_data)
                    
                    if data:
                        gsheet_name = f"{file_type}_{sheet_name.replace(' ', '_')}"
                        if len(gsheet_name) > 100:
                            gsheet_name = gsheet_name[:97] + "..."
                        
                        all_sheets_data[gsheet_name] = {
                            'original_name': sheet_name,
                            'data': data
                        }
                    
                    pbar.update(1)
            
            print(f"📤 Google Sheets에 업로드 중... (총 {len(all_sheets_data)}개 시트)")
            self._batch_upload_to_google_sheets(all_sheets_data, rcept_no)
            
        except Exception as e:
            print(f"❌ Excel 처리 실패: {str(e)}")
            self.results['xbrl']['failed_uploads'].append(file_path)

    def _batch_upload_to_google_sheets(self, all_sheets_data, rcept_no):
        """여러 시트를 배치로 Google Sheets에 업로드"""
        try:
            existing_sheets = [ws.title for ws in self.workbook.worksheets()]
            
            sheets_to_create = []
            sheets_to_update = []
            
            for gsheet_name in all_sheets_data:
                if gsheet_name in existing_sheets:
                    sheets_to_update.append(gsheet_name)
                else:
                    sheets_to_create.append(gsheet_name)
            
            # 새 시트 생성
            if sheets_to_create:
                print(f"🆕 새 시트 {len(sheets_to_create)}개 생성 중...")
                
                batch_size = 5
                for i in range(0, len(sheets_to_create), batch_size):
                    batch = sheets_to_create[i:i + batch_size]
                    
                    for sheet_name in batch:
                        try:
                            data = all_sheets_data[sheet_name]['data']
                            rows = max(1000, len(data) + 100)
                            cols = max(26, len(data[0]) + 5) if data else 26
                            self.workbook.add_worksheet(sheet_name, rows, cols)
                        except Exception as e:
                            print(f"⚠️ 시트 생성 실패 {sheet_name}: {str(e)}")
                    
                    time.sleep(3)
            
            # 기존 시트 클리어
            if sheets_to_update:
                print(f"🧹 기존 시트 {len(sheets_to_update)}개 초기화 중...")
                for sheet_name in sheets_to_update:
                    try:
                        worksheet = self.workbook.worksheet(sheet_name)
                        worksheet.clear()
                        time.sleep(1)
                    except Exception as e:
                        print(f"⚠️ 시트 초기화 실패 {sheet_name}: {str(e)}")
            
            # 데이터 업로드
            print(f"📝 데이터 업로드 중...")
            
            upload_count = 0
            total_sheets = len(all_sheets_data)
            
            with tqdm(total=total_sheets, desc="시트 업로드", unit="시트") as pbar:
                for gsheet_name, sheet_info in all_sheets_data.items():
                    try:
                        worksheet = self.workbook.worksheet(gsheet_name)
                        
                        header = [
                            [f"업데이트: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"],
                            [f"보고서: {rcept_no}"],
                            [f"원본 시트: {sheet_info['original_name']}"],
                            []
                        ]
                        
                        all_data = header + sheet_info['data']
                        
                        if len(all_data) > 0:
                            end_row = len(all_data)
                            end_col = max(len(row) for row in all_data) if all_data else 1
                            end_col_letter = self._get_column_letter(end_col - 1)
                            
                            range_name = f'A1:{end_col_letter}{end_row}'
                            worksheet.update(values=all_data, range_name=range_name)
                        
                        self.results['xbrl']['uploaded_sheets'].append(gsheet_name)
                        upload_count += 1
                        
                        if upload_count % 10 == 0:
                            print(f"  💤 API 제한 회피를 위해 10초 대기 중...")
                            time.sleep(10)
                        else:
                            time.sleep(2)
                        
                    except Exception as e:
                        print(f"❌ 시트 업로드 실패 '{gsheet_name}': {str(e)}")
                        self.results['xbrl']['failed_uploads'].append(gsheet_name)
                        
                        if "429" in str(e):
                            print(f"  ⏳ API 할당량 초과. 30초 대기 중...")
                            time.sleep(30)
                    
                    pbar.update(1)
            
            print(f"✅ 업로드 완료: 성공 {upload_count}/{total_sheets}개")
            
        except Exception as e:
            print(f"❌ 배치 업로드 실패: {str(e)}")

    def _update_xbrl_archive_for_current_report(self):
        """현재 보고서의 XBRL Archive 업데이트"""
        print("📊 현재 문서 XBRL Archive 업데이트 중...")
        
        try:
            if 'financial' in self.results['xbrl']['excel_files']:
                print("📈 재무제표 Archive 업데이트...")
                self._update_single_xbrl_archive('Dart_Archive_XBRL_재무제표', 
                                               self.results['xbrl']['excel_files']['financial'], 
                                               'financial')
            
            if 'notes' in self.results['xbrl']['excel_files']:
                print("📝 주석 Archive 업데이트...")
                
                self._update_single_xbrl_archive('Dart_Archive_XBRL_주석_연결', 
                                               self.results['xbrl']['excel_files']['notes'], 
                                               'notes_consolidated')
                
                self._update_single_xbrl_archive('Dart_Archive_XBRL_주석_별도', 
                                               self.results['xbrl']['excel_files']['notes'], 
                                               'notes_standalone')
            
            print("✅ 현재 문서 XBRL Archive 업데이트 완료")
            
        except Exception as e:
            print(f"❌ 현재 문서 XBRL Archive 업데이트 실패: {str(e)}")

    def _update_single_xbrl_archive(self, sheet_name, file_path, file_type):
        """개별 XBRL Archive 시트 업데이트"""
        try:
            # Archive 시트 가져오기 또는 생성
            archive_exists = False
            try:
                archive_sheet = self.workbook.worksheet(sheet_name)
                archive_exists = True
                print(f"📄 기존 {sheet_name} 시트 발견")
            except gspread.exceptions.WorksheetNotFound:
                print(f"🆕 새로운 {sheet_name} 시트 생성")
                time.sleep(2)
                max_rows = 2000 if 'notes' in file_type else 1000
                archive_sheet = self.workbook.add_worksheet(sheet_name, max_rows, 20)
                time.sleep(2)
            
            # 시트가 새로 생성된 경우 헤더 설정
            if not archive_exists:
                header_type = file_type
                if file_type.startswith('notes_'):
                    header_type = 'notes'
                self._setup_xbrl_archive_header(archive_sheet, header_type)
                time.sleep(3)
            
            # 현재 마지막 데이터 열 찾기 (M열부터)
            last_col = self._find_last_data_column(archive_sheet)
            
            # Excel 파일 읽기
            wb = load_workbook(file_path, data_only=True)
            
            # 데이터 추출 및 업데이트
            if file_type == 'financial':
                self._update_xbrl_financial_archive_batch(archive_sheet, wb, last_col)
            elif file_type == 'notes_consolidated':
                self._update_xbrl_notes_archive_batch(archive_sheet, wb, last_col, 'consolidated')
            elif file_type == 'notes_standalone':
                self._update_xbrl_notes_archive_batch(archive_sheet, wb, last_col, 'standalone')
                
        except Exception as e:
            print(f"❌ {sheet_name} 업데이트 실패: {str(e)}")
            
            if "429" in str(e):
                print(f"  ⏳ API 할당량 초과. 60초 대기 중...")
                time.sleep(60)

    def _setup_xbrl_archive_header(self, sheet, file_type):
        """XBRL Archive 시트 헤더 설정"""
        try:
            current_date = datetime.now().strftime('%Y-%m-%d')
            
            number_unit = os.environ.get('NUMBER_UNIT', 'million')
            unit_text = {
                'million': '백만원',
                'hundred_million': '억원',
                'billion': '십억원'
            }.get(number_unit, '백만원')
            
            header_data = []
            
            if file_type == 'financial':
                title_row = ['DART Archive XBRL 재무제표', '', '', '', '', '', '', '', '', f'최종업데이트: {current_date}', '', '계정과목']
            else:
                title_row = ['DART Archive XBRL 재무제표주석', '', '', '', '', '', '', '', '', f'최종업데이트: {current_date}', '', '계정과목']
            header_data.append(title_row)
            
            company_row = [f'회사명: {self.company_name}', '', '', '', '', '', '', '', '', f'단위: {unit_text}', '', '항목명↓']
            header_data.append(company_row)
            
            stock_row = [f'종목코드: {self.corp_code}', '', '', '', '', '', '', '', '', '', '', '']
            header_data.append(stock_row)
            
            for _ in range(3):
                header_data.append(['', '', '', '', '', '', '', '', '', '', '', ''])
            
            end_row = len(header_data)
            range_name = f'A1:L{end_row}'
            
            print(f"  📋 XBRL Archive 기본 헤더 설정: {range_name}")
            sheet.update(values=header_data, range_name=range_name)
            
            print(f"  ✅ XBRL Archive 기본 레이아웃 완료")
            
        except Exception as e:
            print(f"  ❌ XBRL Archive 헤더 설정 실패: {str(e)}")

    def _find_last_data_column(self, sheet):
        """마지막 데이터 열 찾기 (M열부터 시작)"""
        try:
            row_2_values = sheet.row_values(2)
            
            last_col = 11  # M열 = 12번째 열 (0-based index에서는 11)
            
            for i in range(11, len(row_2_values)):
                if row_2_values[i]:
                    last_col = i
            
            next_col = last_col + 1
            
            if next_col < 11:
                next_col = 11
            
            col_letter = self._get_column_letter(next_col)
            print(f"📍 새 데이터 추가 위치: {col_letter}열 (인덱스: {next_col})")
            
            return next_col
            
        except Exception as e:
            print(f"⚠️ 마지막 열 찾기 실패: {str(e)}")
            return 11

    def _update_xbrl_financial_archive_batch(self, sheet, wb, col_index):
        """XBRL 재무제표 Archive 업데이트"""
        try:
            print(f"  📊 XBRL 재무제표 데이터 추출 중...")
            
            col_letter = self._get_column_letter(col_index)
            print(f"  📍 데이터 입력 위치: {col_letter}열")
            
            # 기존 L열의 계정명 읽어오기
            existing_accounts = set()
            try:
                l_column_values = sheet.col_values(12)
                for idx, account in enumerate(l_column_values[6:], start=7):
                    if account and account.strip():
                        existing_accounts.add(account.strip())
                
                print(f"  📋 기존 계정명 {len(existing_accounts)}개 발견")
            except Exception as e:
                print(f"  ⚠️ 기존 계정명 읽기 실패: {str(e)}")
            
            # 헤더 정보 업데이트
            report_date = datetime.now().strftime('%Y-%m-%d')
            quarter_info = self._get_quarter_info()
            
            # 모든 재무 데이터를 메모리에서 준비
            all_account_data, all_value_data = self._prepare_financial_data_for_batch_update(wb)
            
            # 신규 계정명 추적
            new_accounts = []
            for idx, account_row in enumerate(all_account_data):
                if account_row and account_row[0]:
                    account_name = account_row[0]
                    if (not account_name.startswith('[') and 
                        not account_name.startswith('===') and
                        account_name not in existing_accounts):
                        new_accounts.append((idx, account_name))
            
            if new_accounts:
                print(f"  🆕 신규 계정명 {len(new_accounts)}개 발견")
            
            # 배치 업데이트
            print(f"  🚀 대용량 배치 업데이트 시작...")
            
            # 헤더 정보
            header_range = f'{col_letter}1:{col_letter}2'
            header_data = [[quarter_info], [report_date]]
            sheet.update(values=header_data, range_name=header_range)
            print(f"    ✅ 헤더 정보 업데이트 완료")
            
            # L열 계정명
            if all_account_data:
                account_range = f'L7:L{6 + len(all_account_data)}'
                sheet.update(values=all_account_data, range_name=account_range)
                print(f"    ✅ L열 계정명 업데이트 완료")
            
            time.sleep(2)
            
            # M열 값
            if all_value_data:
                value_range = f'{col_letter}7:{col_letter}{6 + len(all_value_data)}'
                sheet.update(values=all_value_data, range_name=value_range)
                print(f"    ✅ {col_letter}열 값 업데이트 완료")
            
            print(f"  ✅ XBRL 재무제표 Archive 배치 업데이트 완료")
            
        except Exception as e:
            print(f"❌ XBRL 재무제표 Archive 업데이트 실패: {str(e)}")

    def _prepare_financial_data_for_batch_update(self, wb):
        """재무 데이터를 배치 업데이트용으로 준비"""
        try:
            print(f"  🔄 배치 업데이트용 데이터 준비 중...")
            
            all_account_data = []
            all_value_data = []
            
            # D로 시작하는 시트 처리
            d_sheets = [name for name in wb.sheetnames if name.startswith('D')]
            print(f"  📋 D로 시작하는 시트 {len(d_sheets)}개 발견")
            
            for sheet_name in sorted(d_sheets):
                worksheet = wb[sheet_name]
                
                # 시트 제목 찾기
                sheet_title = self._find_sheet_title(worksheet) or sheet_name
                
                # 연결/별도 구분
                sheet_type = ""
                if '연결' in sheet_title or sheet_name.endswith('0'):
                    sheet_type = "[연결]"
                elif '별도' in sheet_title or sheet_name.endswith('5'):
                    sheet_type = "[별도]"
                else:
                    sheet_type = "[기타]"
                
                # 재무제표 종류 판단
                fs_type = ""
                if '재무상태표' in sheet_title:
                    fs_type = "재무상태표"
                elif '손익계산서' in sheet_title or '포괄손익' in sheet_title:
                    fs_type = "손익계산서"
                elif '현금흐름표' in sheet_title:
                    fs_type = "현금흐름표"
                elif '자본변동표' in sheet_title:
                    fs_type = "자본변동표"
                else:
                    continue
                
                # 시트명 헤더 추가
                header_text = f"{sheet_type} {fs_type} ({sheet_name})"
                all_account_data.append([header_text])
                all_value_data.append([''])
                
                # 데이터 추출
                data_count = 0
                for row_idx in range(1, min(worksheet.max_row + 1, 500)):
                    row = list(worksheet.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))[0]
                    
                    if not row or len(row) < 2:
                        continue
                    
                    # A열: 계정명
                    account_name = str(row[0]).strip() if row[0] else ''
                    
                    if (not account_name or 
                        len(account_name) < 2 or 
                        account_name.startswith('[') or
                        account_name.startswith('(단위')):
                        continue
                    
                    # B열: 값
                    value = None
                    if row[1] is not None:
                        if isinstance(row[1], (int, float)):
                            value = row[1]
                        elif isinstance(row[1], str):
                            try:
                                clean_str = str(row[1]).replace(',', '').replace('(', '-').replace(')', '').strip()
                                if clean_str and clean_str != '-':
                                    value = float(clean_str)
                            except:
                                pass
                    
                    all_account_data.append([account_name])
                    all_value_data.append([self._format_number_for_archive(value) if value else ''])
                    data_count += 1
                
                if data_count > 0:
                    print(f"    ✅ {sheet_name}: {data_count}개 항목 추가")
                    all_account_data.append([''])
                    all_value_data.append([''])
            
            return all_account_data, all_value_data
            
        except Exception as e:
            print(f"  ❌ 배치 데이터 준비 실패: {str(e)}")
            return [], []

    def _find_sheet_title(self, worksheet):
        """시트 제목 찾기"""
        try:
            for row in worksheet.iter_rows(min_row=1, max_row=10, values_only=True):
                for cell in row:
                    if cell and isinstance(cell, str):
                        if any(keyword in str(cell) for keyword in ['재무상태표', '손익계산서', '현금흐름표', '자본변동표', '포괄손익']):
                            return str(cell).strip()
            return None
        except:
            return None

    def _update_xbrl_notes_archive_batch(self, sheet, wb, col_index, notes_type='consolidated'):
        """XBRL 재무제표주석 Archive 업데이트"""
        try:
            print(f"  📝 XBRL 주석 데이터 분석 중... ({notes_type})")
            
            col_letter = self._get_column_letter(col_index)
            print(f"  📍 데이터 입력 위치: {col_letter}열")
            
            # 헤더 정보 업데이트
            report_date = datetime.now().strftime('%Y-%m-%d')
            quarter_info = self._get_quarter_info()
            
            # 모든 주석 데이터를 메모리에서 준비
            all_notes_account_data, all_notes_value_data = self._prepare_notes_data_for_batch_update(wb, notes_type)
            
            # 배치 업데이트
            print(f"  🚀 주석 배치 업데이트 시작...")
            
            # 헤더 정보
            header_range = f'{col_letter}1:{col_letter}2'
            header_data = [[quarter_info], [report_date]]
            sheet.update(values=header_data, range_name=header_range)
            print(f"    ✅ 헤더 정보 업데이트 완료")
            
            # L열 주석 항목명
            if all_notes_account_data:
                account_range = f'L7:L{6 + len(all_notes_account_data)}'
                sheet.update(values=all_notes_account_data, range_name=account_range)
                print(f"    ✅ L열 주석 항목 업데이트 완료")
            
            time.sleep(2)
            
            # M열 주석 값
            if all_notes_value_data:
                value_range = f'{col_letter}7:{col_letter}{6 + len(all_notes_value_data)}'
                sheet.update(values=all_notes_value_data, range_name=value_range)
                print(f"    ✅ {col_letter}열 주석 값 업데이트 완료")
            
            print(f"  ✅ XBRL 주석 Archive 배치 업데이트 완료")
            
        except Exception as e:
            print(f"❌ XBRL 주석 Archive 업데이트 실패: {str(e)}")

    def _prepare_notes_data_for_batch_update(self, wb, notes_type):
        """주석 데이터를 배치 업데이트용으로 준비 (개선된 로직)"""
        try:
            print(f"  🔄 주석 배치 업데이트용 데이터 준비 중... ({notes_type})")
            
            # 전체 시트 목록 출력
            print(f"    📋 전체 시트 목록: {wb.sheetnames}")
            
            # 개선된 주석 시트 찾기 로직
            target_sheets = self._find_notes_sheets(wb, notes_type)
            
            print(f"    📄 {notes_type} 주석 시트 {len(target_sheets)}개 발견: {target_sheets}")
            
            # 전체 데이터를 하나의 배열로 통합
            all_notes_account_data = []
            all_notes_value_data = []
            
            # 각 주석 시트의 데이터 추출 및 배치
            for sheet_name in sorted(target_sheets):
                sheet_data = self._extract_notes_sheet_data_improved(wb[sheet_name], sheet_name)
                if sheet_data:
                    # 시트 제목 추가
                    all_notes_account_data.append([f"===== {sheet_data['title']} ====="])
                    all_notes_value_data.append([''])
                    
                    # 각 항목들 배치
                    for item in sheet_data['items']:
                        if item.get('is_category'):
                            display_name = item['name']
                        elif 'display_name' in item:
                            display_name = item['display_name']
                        else:
                            original_name = item.get('original_name', item['name'])
                            indent_level = item.get('indent_level', 0)
                            
                            if indent_level > 0:
                                display_name = "  " * indent_level + "└ " + original_name
                            else:
                                display_name = original_name
                        
                        all_notes_account_data.append([display_name])
                        all_notes_value_data.append([item['formatted_value']])
                    
                    # 구분을 위한 빈 행 추가
                    all_notes_account_data.append([''])
                    all_notes_value_data.append([''])
            
            # 통계 출력
            total_items = len([row for row in all_notes_account_data if row[0] and not row[0].startswith('=')])
            print(f"    📊 총 주석 항목: {total_items}개")
            
            return all_notes_account_data, all_notes_value_data
            
        except Exception as e:
            print(f"  ❌ 주석 배치 데이터 준비 실패: {str(e)}")
            return [], []

    def _find_notes_sheets(self, wb, notes_type):
        """주석 시트를 찾는 개선된 로직 (D8/U8 규칙 적용)"""
        target_sheets = []
        
        print(f"    🔍 {notes_type} 주석 시트 검색 중...")
        
        for sheet_name in wb.sheetnames:
            if sheet_name in ['Index', '공시기본정보']:
                continue
            
            is_target_sheet = False
            
            # 주석 시트 명명 규칙 체크: D8/U8로 시작하고 연결(0)/별도(5)로 끝남
            if notes_type == 'consolidated':
                # 연결: D8xxx0 또는 U8xxx0
                if (sheet_name.startswith('D8') or sheet_name.startswith('U8')) and sheet_name.endswith('0'):
                    is_target_sheet = True
                    print(f"      ✅ 연결 주석 시트 발견: {sheet_name}")
            else:  # standalone
                # 별도: D8xxx5 또는 U8xxx5
                if (sheet_name.startswith('D8') or sheet_name.startswith('U8')) and sheet_name.endswith('5'):
                    is_target_sheet = True
                    print(f"      ✅ 별도 주석 시트 발견: {sheet_name}")
            
            # 추가: 내용 기반 체크 (위 규칙에 맞지 않지만 주석일 가능성이 있는 시트)
            if not is_target_sheet:
                # 시트명에 '주석'이 명시적으로 포함된 경우
                if '주석' in sheet_name or 'Notes' in sheet_name or 'Note' in sheet_name:
                    worksheet = wb[sheet_name]
                    sheet_title = self._get_sheet_title(worksheet)
                    
                    if notes_type == 'consolidated':
                        if '연결' in sheet_title or ('별도' not in sheet_title and not sheet_name.endswith('5')):
                            is_target_sheet = True
                            print(f"      ✅ 내용 기반 연결 주석 시트: {sheet_name}")
                    else:
                        if '별도' in sheet_title or sheet_name.endswith('5'):
                            is_target_sheet = True
                            print(f"      ✅ 내용 기반 별도 주석 시트: {sheet_name}")
            
            if is_target_sheet:
                target_sheets.append(sheet_name)
        
        return target_sheets

    def _extract_notes_sheet_data_improved(self, worksheet, sheet_name):
        """개별 주석 시트에서 데이터 추출 (긴 텍스트 처리 개선)"""
        try:
            sheet_data = {
                'title': sheet_name,
                'items': []
            }
            
            print(f"\n      🔍 {sheet_name} 주석 시트 분석 중...")
            
            # 전체 시트 스캔
            max_row = min(worksheet.max_row, 1000)
            max_col = min(worksheet.max_column, 20)
            
            # 모든 셀 데이터를 메모리에 로드
            all_data = []
            for row in worksheet.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col, values_only=True):
                all_data.append(list(row))
            
            print(f"      📊 시트 크기: {len(all_data)}행 x {max_col}열")
            
            # 현재 중분류
            current_category = ""
            current_subcategory = ""
            last_item = None  # 마지막으로 추가한 항목
            
            for row_idx, row in enumerate(all_data):
                if not row or not any(row):  # 빈 행 건너뛰기
                    continue
                
                # 첫 번째 비어있지 않은 셀의 위치와 내용 찾기
                first_text = None
                first_col = -1
                
                for col_idx, cell in enumerate(row):
                    if cell and str(cell).strip():
                        first_text = str(cell).strip()
                        first_col = col_idx
                        break
                
                if not first_text or len(first_text) < 2:
                    continue
                
                # 제외할 패턴 (단위 표시 등)
                if any(skip in first_text for skip in ['(단위', '단위:', 'Index', 'Sheet']):
                    continue
                
                # 대괄호로 둘러싸인 텍스트는 분류명으로 처리
                if first_text.startswith('[') and first_text.endswith(']'):
                    category_name = first_text[1:-1]  # 대괄호 제거
                    current_category = category_name
                    current_subcategory = ""  # 새 중분류시 하위분류 초기화
                    
                    sheet_data['items'].append({
                        'name': f"[중분류] {category_name}",
                        'value': None,
                        'formatted_value': '',
                        'category': category_name,
                        'is_category': True,
                        'original_name': first_text
                    })
                    last_item = None  # 카테고리 변경시 리셋
                    continue
                
                # 긴 텍스트 판별 (50자 이상)
                is_long_text = len(first_text) > 50
                
                # 긴 텍스트이고, 바로 이전에 짧은 항목명이 있는 경우
                if is_long_text and last_item and not last_item.get('is_category'):
                    # 이전 항목의 값으로 처리
                    if last_item.get('value'):
                        # 이미 값이 있으면 추가
                        existing_value = str(last_item['value'])
                        last_item['value'] = existing_value + "\n" + first_text
                    else:
                        # 값이 없으면 새로 설정
                        last_item['value'] = first_text
                        last_item['value_type'] = 'text'
                    
                    # formatted_value 업데이트
                    last_item['formatted_value'] = self._format_notes_value(last_item['value'], 'text')
                    continue
                
                # A열이 비어있고 B열(또는 그 이후)에 텍스트가 있는 경우 - 들여쓰기된 항목
                if first_col > 0:
                    # 들여쓰기된 항목으로 처리
                    indent_level = first_col
                    
                    # 긴 텍스트이고 마지막 항목이 있으면 그 항목의 값으로 처리
                    if is_long_text and last_item and not last_item.get('is_category'):
                        if last_item.get('value'):
                            existing_value = str(last_item['value'])
                            last_item['value'] = existing_value + "\n" + ("  " * indent_level) + first_text
                        else:
                            last_item['value'] = ("  " * indent_level) + first_text
                            last_item['value_type'] = 'text'
                        
                        last_item['formatted_value'] = self._format_notes_value(last_item['value'], 'text')
                        continue
                    
                    # 일반적인 들여쓰기 항목 처리
                    value = None
                    value_type = None
                    
                    # 값 찾기
                    for col_idx in range(first_col + 1, len(row)):
                        if row[col_idx] is not None:
                            value, value_type = self._extract_cell_value(row[col_idx])
                            if value is not None:
                                break
                    
                    # 들여쓰기 표시와 함께 항목 추가
                    display_name = "  " * indent_level + "└ " + first_text
                    unique_name = f"{current_category}_{current_subcategory}_{first_text}" if current_subcategory else f"{current_category}_{first_text}"
                    
                    new_item = {
                        'name': unique_name,
                        'original_name': first_text,
                        'display_name': display_name,
                        'value': value,
                        'formatted_value': self._format_notes_value(value, value_type) if value is not None else '',
                        'category': current_category,
                        'subcategory': current_subcategory,
                        'is_category': False,
                        'row_number': row_idx + 1,
                        'value_type': value_type,
                        'indent_level': indent_level
                    }
                    sheet_data['items'].append(new_item)
                    last_item = new_item
                else:
                    # A열에 있는 항목 (들여쓰기 없음)
                    # 하위 분류일 가능성 체크
                    is_subcategory = False
                    
                    # 다음 행들이 들여쓰기되어 있는지 확인
                    if row_idx + 1 < len(all_data) and not is_long_text:
                        next_rows_indented = 0
                        for check_idx in range(row_idx + 1, min(row_idx + 6, len(all_data))):
                            if check_idx < len(all_data):
                                check_row = all_data[check_idx]
                                # A열이 비어있고 B열 이후에 데이터가 있는지 확인
                                if check_row and (not check_row[0] or not str(check_row[0]).strip()):
                                    for col in range(1, min(5, len(check_row))):
                                        if check_row[col] and str(check_row[col]).strip():
                                            next_rows_indented += 1
                                            break
                        
                        if next_rows_indented >= 2:
                            is_subcategory = True
                            current_subcategory = first_text
                    
                    if is_subcategory:
                        # 하위 분류로 처리
                        new_item = {
                            'name': f"[하위분류] {first_text}",
                            'value': None,
                            'formatted_value': '',
                            'category': current_category,
                            'subcategory': first_text,
                            'is_category': True,
                            'is_subcategory': True,
                            'original_name': first_text
                        }
                        sheet_data['items'].append(new_item)
                        last_item = new_item
                    else:
                        # 일반 항목으로 처리
                        # 값 찾기
                        value = None
                        value_type = None
                        
                        # 같은 행의 다음 열들에서 값 찾기
                        for col_idx in range(first_col + 1, len(row)):
                            if row[col_idx] is not None:
                                value, value_type = self._extract_cell_value(row[col_idx])
                                if value is not None:
                                    break
                        
                        unique_name = f"{current_category}_{current_subcategory}_{first_text}" if current_subcategory else f"{current_category}_{first_text}" if current_category else first_text
                        
                        new_item = {
                            'name': unique_name,
                            'original_name': first_text,
                            'value': value,
                            'formatted_value': self._format_notes_value(value, value_type) if value is not None else '',
                            'category': current_category,
                            'subcategory': current_subcategory,
                            'is_category': False,
                            'row_number': row_idx + 1,
                            'value_type': value_type,
                            'indent_level': 0,
                            'text_length': len(first_text)
                        }
                        sheet_data['items'].append(new_item)
                        last_item = new_item
            
            # 결과 요약
            if sheet_data['items']:
                category_count = len([item for item in sheet_data['items'] if item.get('is_category') and not item.get('is_subcategory')])
                subcategory_count = len([item for item in sheet_data['items'] if item.get('is_subcategory')])
                value_count = len([item for item in sheet_data['items'] if item.get('value') is not None])
                text_count = len([item for item in sheet_data['items'] if item.get('value_type') == 'text'])
                number_count = len([item for item in sheet_data['items'] if item.get('value_type') == 'number'])
                
                print(f"      ✅ 추출 완료: 총 {len(sheet_data['items'])}개 항목")
                print(f"         - 중분류: {category_count}개")
                print(f"         - 하위분류: {subcategory_count}개") 
                print(f"         - 값 있음: {value_count}개 (숫자: {number_count}, 텍스트: {text_count})")
            
            return sheet_data if sheet_data['items'] else None
            
        except Exception as e:
            print(f"      ❌ 주석 시트 {sheet_name} 데이터 추출 실패: {str(e)}")
            import traceback
            traceback.print_exc()
            return None

    def _extract_cell_value(self, cell_value):
        """셀 값에서 실제 값과 타입 추출"""
        if cell_value is None:
            return None, None
            
        # 숫자인 경우
        if isinstance(cell_value, (int, float)):
            return cell_value, 'number'
        
        # 문자열인 경우
        elif isinstance(cell_value, str):
            str_val = str(cell_value).strip()
            if not str_val or str_val == '-':
                return None, None
                
            # 숫자 변환 시도
            try:
                clean_num = str_val.replace(',', '').replace('(', '-').replace(')', '').strip()
                if clean_num and clean_num != '-' and clean_num.replace('-', '').replace('.', '').isdigit():
                    return float(clean_num), 'number'
            except:
                pass
            
            # 텍스트로 처리
            if len(str_val) >= 2:
                return str_val, 'text'
        
        return None, None

    def _format_notes_value(self, value, value_type=None):
        """주석 값 포맷팅"""
        try:
            if value is None:
                return ''
            
            # 텍스트인 경우
            if value_type == 'text' or isinstance(value, str):
                text_value = str(value).strip()
                if len(text_value) > 100:
                    return text_value[:97] + "..."
                else:
                    return text_value
            
            # 숫자인 경우
            elif isinstance(value, (int, float)):
                number_unit = os.environ.get('NUMBER_UNIT', 'million')
                
                if number_unit == 'million':
                    if abs(value) >= 1000000:
                        converted_value = value / 1000000
                        return f"{converted_value:.1f}백만원"
                    else:
                        return f"{value:,.0f}"
                elif number_unit == 'hundred_million':
                    if abs(value) >= 100000000:
                        converted_value = value / 100000000
                        return f"{converted_value:.2f}억원"
                    elif abs(value) >= 1000000:
                        million_value = value / 1000000
                        return f"{million_value:.1f}백만원"
                    else:
                        return f"{value:,.0f}"
                elif number_unit == 'billion':
                    if abs(value) >= 1000000000:
                        converted_value = value / 1000000000
                        return f"{converted_value:.2f}십억원"
                    elif abs(value) >= 100000000:
                        hundred_million_value = value / 100000000
                        return f"{hundred_million_value:.1f}억원"
                    else:
                        return f"{value:,.0f}"
                else:
                    if abs(value) >= 1000000:
                        converted_value = value / 1000000
                        return f"{converted_value:.1f}백만원"
                    else:
                        return f"{value:,.0f}"
            else:
                return str(value)
                
        except Exception as e:
            print(f"    ⚠️ 주석 값 포맷팅 오류 ({value}): {str(e)}")
            return str(value) if value else ''

    def _format_number_for_archive(self, value):
        """Archive용 숫자 포맷팅"""
        try:
            if not value:
                return ''
            
            num = self._clean_number(value)
            if num is None:
                return ''
            
            number_unit = os.environ.get('NUMBER_UNIT', 'million')
            
            if number_unit == 'million':
                unit_value = num / 1000000
            elif number_unit == 'hundred_million':
                unit_value = num / 100000000
            elif number_unit == 'billion':
                unit_value = num / 1000000000
            else:
                unit_value = num / 1000000
            
            if abs(unit_value) >= 1000:
                formatted = f"{unit_value:.0f}"
            elif abs(unit_value) >= 100:
                formatted = f"{unit_value:.1f}"
            else:
                formatted = f"{unit_value:.2f}"
            
            return formatted
                
        except Exception as e:
            print(f"    ⚠️ 숫자 포맷팅 오류 ({value}): {str(e)}")
            return str(value)

    def _clean_number(self, value):
        """숫자 값 정제"""
        try:
            if isinstance(value, (int, float)):
                return float(value)
            
            str_val = str(value).replace(',', '').replace('(', '-').replace(')', '').strip()
            if not str_val or str_val == '-':
                return None
            return float(str_val)
        except:
            return None

    def _get_quarter_info(self):
        """보고서 기준 분기 정보 반환"""
        try:
            if self.current_report is not None:
                report_name = self.current_report.get('report_nm', '')
                
                if report_name:
                    print(f"  📅 보고서 분석: {report_name}")
                    
                    if '1분기' in str(report_name):
                        current_year = datetime.now().year
                        quarter_text = f"1Q{str(current_year)[2:]}"
                        return quarter_text
                    elif '반기' in str(report_name) or '2분기' in str(report_name):
                        current_year = datetime.now().year
                        quarter_text = f"2Q{str(current_year)[2:]}"
                        return quarter_text
                    elif '3분기' in str(report_name):
                        current_year = datetime.now().year
                        quarter_text = f"3Q{str(current_year)[2:]}"
                        return quarter_text
                    
                    # 날짜 패턴 매칭
                    date_pattern1 = re.search(r'\((\d{4})\.(\d{2})\)', str(report_name))
                    date_pattern2 = re.search(r'(\d{4})년\s*(\d{1,2})월', str(report_name))
                    
                    year, month = None, None
                    
                    if date_pattern1:
                        year, month = date_pattern1.groups()
                        month = int(month)
                    elif date_pattern2:
                        year, month = date_pattern2.groups()
                        month = int(month)
                    
                    if year and month:
                        if month <= 3:
                            quarter = 1
                        elif month <= 6:
                            quarter = 2
                        elif month <= 9:
                            quarter = 3
                        else:
                            quarter = 4
                        
                        quarter_text = f"{quarter}Q{year[2:]}"
                        return quarter_text
        
        except Exception as e:
            print(f"    ⚠️ 분기 정보 추출 중 오류: {str(e)}")
        
        # 기본값: 현재 날짜 기준
        now = datetime.now()
        quarter = (now.month - 1) // 3 + 1
        year = str(now.year)[2:]
        default_quarter = f"{quarter}Q{year}"
        return default_quarter

    def _get_sheet_title(self, worksheet):
        """시트의 제목 찾기"""
        try:
            for row_idx in range(1, min(11, worksheet.max_row + 1)):
                for col_idx in range(1, min(4, worksheet.max_column + 1)):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    if cell.value and isinstance(cell.value, str):
                        value = str(cell.value).strip()
                        if len(value) > 5 and ('재무상태표' in value or '손익계산서' in value or 
                                               '현금흐름표' in value or '자본변동표' in value or
                                               '포괄손익' in value or '주석' in value):
                            return value
            return ""
        except:
            return ""

    def _get_column_letter(self, col_index):
        """컬럼 인덱스를 문자로 변환 (0-based)"""
        result = ""
        num = col_index + 1
        while num > 0:
            num, remainder = divmod(num - 1, 26)
            result = chr(65 + remainder) + result
        return result

    def _cleanup_current_downloads(self):
        """현재 문서 다운로드 파일 정리"""
        try:
            if os.path.exists(self.download_dir):
                for file in os.listdir(self.download_dir):
                    file_path = os.path.join(self.download_dir, file)
                    if os.path.isfile(file_path):
                        os.remove(file_path)
                print("🧹 현재 문서 파일 정리 완료")
            
            # Excel 파일 경로 초기화
            self.results['xbrl']['excel_files'] = {}
            
        except Exception as e:
            print(f"⚠️ 현재 문서 파일 정리 실패: {str(e)}")

    def _send_telegram_message(self, message):
        """텔레그램 메시지 전송"""
        try:
            if self.telegram_bot_token and self.telegram_channel_id:
                import requests
                url = f"https://api.telegram.org/bot{self.telegram_bot_token}/sendMessage"
                data = {
                    "chat_id": self.telegram_channel_id,
                    "text": message,
                    "parse_mode": "HTML"
                }
                requests.post(url, data=data)
                print("📱 텔레그램 메시지 전송 완료")
        except Exception as e:
            print(f"📱 텔레그램 메시지 전송 실패: {str(e)}")

    def _cleanup_downloads(self):
        """다운로드 폴더 정리"""
        try:
            if os.path.exists(self.download_dir) and self.results.get('xbrl', {}).get('excel_files'):
                for file in os.listdir(self.download_dir):
                    file_path = os.path.join(self.download_dir, file)
                    if file_path not in self.results['xbrl']['downloaded_files']:
                        os.remove(file_path)
                
                if os.environ.get('DELETE_AFTER_ARCHIVE', 'true').lower() == 'true':
                    shutil.rmtree(self.download_dir)
                    print("🧹 다운로드 폴더 정리 완료")
                else:
                    print("📁 다운로드 파일 보존 중")
        except Exception as e:
            print(f"⚠️ 다운로드 폴더 정리 실패: {str(e)}")

    def _print_summary(self):
        """처리 결과 요약"""
        print("\n" + "="*50)
        print("📊 처리 결과 요약")
        print("="*50)
        print(f"전체 보고서: {self.results['total_reports']}개")
        print(f"XBRL 다운로드 성공: {len(self.results['xbrl']['downloaded_files'])}개")
        print(f"XBRL 업로드된 시트: {len(self.results['xbrl']['uploaded_sheets'])}개")
        print(f"XBRL 다운로드 실패: {len(self.results['xbrl']['failed_downloads'])}개")
        print(f"XBRL 업로드 실패: {len(self.results['xbrl']['failed_uploads'])}개")
        print(f"HTML 처리된 시트: {len(self.results['html']['processed_sheets'])}개")
        print(f"HTML 실패: {len(self.results['html']['failed_sheets'])}개")
        
        if self.telegram_bot_token and self.telegram_channel_id:
            self._send_telegram_summary()

    def _send_telegram_summary(self):
        """텔레그램 요약 메시지 전송"""
        try:
            import requests
            
            message = (
                f"📊 DART 통합 업데이트 완료\n\n"
                f"• 종목: {self.company_name} ({self.corp_code})\n"
                f"• 처리 시간: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n"
                f"• 전체 보고서: {self.results['total_reports']}개\n"
                f"• XBRL 다운로드: {len(self.results['xbrl']['downloaded_files'])}개\n"
                f"• XBRL 업로드: {len(self.results['xbrl']['uploaded_sheets'])}개\n"
                f"• HTML 처리: {len(self.results['html']['processed_sheets'])}개"
            )
            
            url = f"https://api.telegram.org/bot{self.telegram_bot_token}/sendMessage"
            data = {
                "chat_id": self.telegram_channel_id,
                "text": message,
                "parse_mode": "HTML"
            }
            requests.post(url, data=data)
            print("📱 텔레그램 메시지 전송 완료")
            
        except Exception as e:
            print(f"📱 텔레그램 메시지 전송 실패: {str(e)}")


def load_company_config():
    """회사 설정 로드"""
    corp_code = os.environ.get('COMPANY_CORP_CODE', '307950')
    company_name = os.environ.get('COMPANY_NAME', '현대오토에버')
    spreadsheet_var = os.environ.get('COMPANY_SPREADSHEET_VAR', 'AUTOEVER_SPREADSHEET_ID')
    
    return {
        'corp_code': corp_code,
        'company_name': company_name,
        'spreadsheet_var': spreadsheet_var
    }


def main():
    """메인 실행 함수"""
    try:
        print("🔧 Playwright 브라우저 설치 확인...")
        os.system("playwright install chromium")
        
        company_config = load_company_config()
        
        print(f"🤖 DART 통합 업데이터 시스템")
        print(f"🏢 대상 기업: {company_config['company_name']} ({company_config['corp_code']})")
        
        updater = DartDualUpdater(company_config)
        updater.run()
        
        print("\n✅ 모든 작업이 완료되었습니다!")
        
    except Exception as e:
        print(f"\n❌ 오류 발생: {str(e)}")
        import traceback
        traceback.print_exc()
        raise


if __name__ == "__main__":
    main()
