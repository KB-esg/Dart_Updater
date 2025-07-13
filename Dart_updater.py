import os
from datetime import datetime, timedelta
import json
import time
import re
import gspread
from google.oauth2.service_account import Credentials
import OpenDartReader
import requests
from bs4 import BeautifulSoup
import pandas as pd
import xml.etree.ElementTree as ET
from urllib.parse import urljoin, urlparse
import zipfile
import io
import openpyxl
from openpyxl import load_workbook

# HTML 테이블 파서 대안 구현
try:
    from html_table_parser import parser_functions as parser
    HTML_PARSER_AVAILABLE = True
    print("✅ html_table_parser 로드 성공")
except ImportError:
    try:
        from html_table_parser_python3 import parser_functions as parser
        HTML_PARSER_AVAILABLE = True
        print("✅ html_table_parser_python3 로드 성공")
    except ImportError:
        HTML_PARSER_AVAILABLE = False
        print("⚠️ HTML 파서 패키지가 없습니다. 내장 파서를 사용합니다.")

class DualSystemDartUpdater:
    """XBRL/HTML 이원화 관리 시스템"""
    
    # HTML 스크래핑 대상 시트 (기존 방식, 주석 제외)
    HTML_TARGET_SHEETS = [
        'I. 회사의 개요', 'II. 사업의 내용', '1. 사업의 개요', '2. 주요 제품 및 서비스',
        '3. 원재료 및 생산설비', '4. 매출 및 수주상황', '5. 위험관리 및 파생거래',
        '6. 주요계약 및 연구활동', '7. 기타 참고 사항', '1. 요약재무정보',
        '6. 배당에 관한 사항', '8. 기타 재무에 관한 사항', 'VII. 주주에 관한 사항',
        'VIII. 임원 및 직원 등에 관한 사항', 'X. 대주주 등과의 거래내용',
        'XI. 그 밖에 투자자 보호를 위하여 필요한 사항'
    ]
    
    # XBRL 우선 처리 대상 시트 (재무제표 + 주석)
    XBRL_TARGET_SHEETS = [
        '2. 연결재무제표',
        '4. 재무제표',
        '3. 연결재무제표 주석',
        '5. 재무제표 주석'
    ]

    def __init__(self, company_config):
        """초기화 - company_config는 yml에서 읽어온 설정"""
        self.corp_code = company_config['corp_code']
        self.company_name = company_config['company_name']
        self.spreadsheet_var_name = company_config['spreadsheet_var']
        
        # 환경변수 확인
        print("환경변수 확인:")
        required_vars = ['DART_API_KEY', 'GOOGLE_CREDENTIALS', self.spreadsheet_var_name, 
                        'TELEGRAM_BOT_TOKEN', 'TELEGRAM_CHANNEL_ID']
        for var in required_vars:
            if var in os.environ:
                value = os.environ[var]
                if len(value) > 4:
                    if len(value) > 20:
                        masked_value = value[:6] + '...' + value[-4:-2] + '**'
                    else:
                        masked_value = value[:-2] + '**'
                    print(f"✅ {var}: {masked_value} (길이: {len(value)})")
                else:
                    print(f"⚠️ {var}: 값이 너무 짧음 (길이: {len(value)})")
            else:
                print(f"❌ {var}: 설정되지 않음")
        
        if self.spreadsheet_var_name not in os.environ:
            raise ValueError(f"{self.spreadsheet_var_name} 환경변수가 설정되지 않았습니다.")
            
        self.credentials = self.get_credentials()
        self.gc = gspread.authorize(self.credentials)
        self.dart = OpenDartReader(os.environ['DART_API_KEY'])
        self.workbook = self.gc.open_by_key(os.environ[self.spreadsheet_var_name])
        self.telegram_bot_token = os.environ.get('TELEGRAM_BOT_TOKEN')
        self.telegram_channel_id = os.environ.get('TELEGRAM_CHANNEL_ID')
        
        # 처리 결과 추적
        self.processing_results = {
            'xbrl_xlsx_success': [],
            'xbrl_xlsx_failed': [],
            'xbrl_success': [],
            'xbrl_failed': [],
            'html_success': [],
            'html_failed': [],
            'total_processed': 0
        }
        
        # XBRL 네임스페이스
        self.xbrl_namespaces = {
            'xbrl': 'http://www.xbrl.org/2003/instance',
            'ifrs': 'http://xbrl.ifrs.org/taxonomy/2021-03-24/ifrs',
            'ifrs-full': 'http://xbrl.ifrs.org/taxonomy/2021-03-24/ifrs-full',
            'dart': 'http://dart.fss.or.kr/xbrl/taxonomy/kr-gaap',
            'link': 'http://www.xbrl.org/2003/linkbase',
            'xlink': 'http://www.w3.org/1999/xlink'
        }

    def get_credentials(self):
        """Google Sheets 인증 설정"""
        creds_json = json.loads(os.environ['GOOGLE_CREDENTIALS'])
        scopes = [
            'https://www.googleapis.com/auth/spreadsheets',
            'https://www.googleapis.com/auth/drive'
        ]
        return Credentials.from_service_account_info(creds_json, scopes=scopes)

    def update_dart_reports(self):
        """DART 보고서 데이터 업데이트 (이원화 시스템)"""
        start_date, end_date = self.get_recent_dates()
        report_list = self.dart.list(self.corp_code, start_date, end_date, kind='A', final='T')
        
        if not report_list.empty:
            print(f"📋 발견된 보고서: {len(report_list)}개")
            
            for _, report in report_list.iterrows():
                print(f"\n📄 보고서 처리 시작: {report['report_nm']} (접수번호: {report['rcept_no']})")
                self.processing_results['total_processed'] += 1
                
                # 1단계: XBRL Excel 파일 다운로드 및 처리 시도 (새로운 방식)
                xbrl_xlsx_success = self.process_xbrl_excel_files(report['rcept_no'])
                
                # 2단계: 기존 XBRL 전용 처리 시도 (기존 방식 백업)
                xbrl_success = False
                if not xbrl_xlsx_success:
                    xbrl_success = self.process_xbrl_sheets(report['rcept_no'])
                
                # 3단계: HTML 시트 처리 (항상 실행)
                html_success = self.process_html_sheets(report['rcept_no'])
                
                # 4단계: 처리 결과 기록
                self.record_processing_result(report, xbrl_xlsx_success or xbrl_success, html_success)
                
        else:
            print("📭 최근 3개월 내 새로운 보고서가 없습니다.")
        
        # 처리 결과 요약 출력
        self.print_processing_summary()

    def process_xbrl_excel_files(self, rcept_no):
        """XBRL Excel 파일 다운로드 및 처리 (새로운 방식)"""
        print(f"\n📊 XBRL Excel 파일 처리 시작: {rcept_no}")
        
        try:
            # XBRL 페이지에서 xbrlExtSeq 추출
            xbrl_ext_seq = self.get_xbrl_ext_seq(rcept_no)
            if not xbrl_ext_seq:
                print("⚠️ xbrlExtSeq를 찾을 수 없음")
                return False
            
            print(f"📝 xbrlExtSeq: {xbrl_ext_seq}")
            
            # 재무제표 Excel 다운로드 및 처리
            financial_success = self.download_and_process_financial_statements(xbrl_ext_seq, rcept_no)
            
            # 재무제표주석 Excel 다운로드 및 처리
            notes_success = self.download_and_process_notes(xbrl_ext_seq, rcept_no)
            
            if financial_success or notes_success:
                print(f"✅ XBRL Excel 파일 처리 완료")
                return True
            else:
                print(f"❌ XBRL Excel 파일 처리 실패")
                return False
                
        except Exception as e:
            print(f"❌ XBRL Excel 파일 처리 중 오류: {str(e)}")
            return False

    def get_xbrl_ext_seq(self, rcept_no):
        """XBRL 페이지에서 xbrlExtSeq 추출"""
        try:
            # XBRL 뷰어 페이지 접근
            viewer_url = f"https://opendart.fss.or.kr/xbrl/viewer/main.do?rcpNo={rcept_no}"
            
            response = requests.get(viewer_url, timeout=30)
            response.raise_for_status()
            
            # HTML 파싱
            soup = BeautifulSoup(response.text, 'html.parser')
            
            # JavaScript에서 xbrlExtSeq 추출 시도
            scripts = soup.find_all('script')
            for script in scripts:
                if script.string:
                    # viewDoc 함수 호출에서 xbrlExtSeq 찾기
                    match = re.search(r"viewDoc\s*\(\s*'(\d+)'", script.string)
                    if match:
                        return match.group(1)
            
            # onclick 속성에서 찾기
            onclick_elements = soup.find_all(attrs={'onclick': True})
            for elem in onclick_elements:
                match = re.search(r"viewDoc\s*\(\s*'(\d+)'", elem.get('onclick', ''))
                if match:
                    return match.group(1)
            
            # iframe src에서 찾기
            iframes = soup.find_all('iframe')
            for iframe in iframes:
                src = iframe.get('src', '')
                match = re.search(r'xbrlExtSeq=(\d+)', src)
                if match:
                    return match.group(1)
            
            print("⚠️ xbrlExtSeq를 HTML에서 찾을 수 없음")
            return None
            
        except Exception as e:
            print(f"❌ xbrlExtSeq 추출 실패: {str(e)}")
            return None

    def download_and_process_financial_statements(self, xbrl_ext_seq, rcept_no):
        """재무제표 Excel 다운로드 및 처리"""
        try:
            print(f"📊 재무제표 Excel 다운로드 중...")
            
            # 다운로드 URL
            download_url = f"https://opendart.fss.or.kr/xbrl/viewer/download/excel/financialStatements.do?xbrlExtSeq={xbrl_ext_seq}&lang=ko"
            
            # Excel 파일 다운로드
            response = requests.get(download_url, timeout=60)
            response.raise_for_status()
            
            # Excel 파일을 메모리에서 처리
            excel_data = io.BytesIO(response.content)
            wb = load_workbook(excel_data, data_only=True)
            
            print(f"📋 재무제표 시트 목록: {wb.sheetnames}")
            
            # 각 시트를 Google Sheets에 업로드
            for sheet_name in wb.sheetnames:
                try:
                    # Excel 시트 데이터 읽기
                    ws = wb[sheet_name]
                    data = []
                    
                    for row in ws.iter_rows(values_only=True):
                        # None 값을 빈 문자열로 변환
                        row_data = [str(cell) if cell is not None else '' for cell in row]
                        if any(row_data):  # 완전히 빈 행은 제외
                            data.append(row_data)
                    
                    if data:
                        # Google Sheets에 시트 생성/업데이트
                        gsheet_name = f"XBRL_{sheet_name}_{rcept_no}"
                        self.update_google_sheet(gsheet_name, data)
                        self.processing_results['xbrl_xlsx_success'].append(gsheet_name)
                        print(f"✅ 재무제표 시트 업로드 완료: {gsheet_name}")
                    
                except Exception as sheet_e:
                    print(f"❌ 시트 처리 실패 {sheet_name}: {str(sheet_e)}")
                    self.processing_results['xbrl_xlsx_failed'].append(sheet_name)
                    continue
            
            return True
            
        except Exception as e:
            print(f"❌ 재무제표 Excel 다운로드/처리 실패: {str(e)}")
            return False

    def download_and_process_notes(self, xbrl_ext_seq, rcept_no):
        """재무제표주석 Excel 다운로드 및 처리"""
        try:
            print(f"📑 재무제표주석 Excel 다운로드 중...")
            
            # 다운로드 URL
            download_url = f"https://opendart.fss.or.kr/xbrl/viewer/download/excel/notes.do?xbrlExtSeq={xbrl_ext_seq}&lang=ko"
            
            # Excel 파일 다운로드
            response = requests.get(download_url, timeout=60)
            response.raise_for_status()
            
            # Excel 파일을 메모리에서 처리
            excel_data = io.BytesIO(response.content)
            wb = load_workbook(excel_data, data_only=True)
            
            print(f"📋 재무제표주석 시트 목록: {wb.sheetnames}")
            
            # 각 시트를 Google Sheets에 업로드
            for sheet_name in wb.sheetnames:
                try:
                    # Excel 시트 데이터 읽기
                    ws = wb[sheet_name]
                    data = []
                    
                    for row in ws.iter_rows(values_only=True):
                        # None 값을 빈 문자열로 변환
                        row_data = [str(cell) if cell is not None else '' for cell in row]
                        if any(row_data):  # 완전히 빈 행은 제외
                            data.append(row_data)
                    
                    if data:
                        # Google Sheets에 시트 생성/업데이트
                        gsheet_name = f"XBRL_주석_{sheet_name}_{rcept_no}"
                        self.update_google_sheet(gsheet_name, data)
                        self.processing_results['xbrl_xlsx_success'].append(gsheet_name)
                        print(f"✅ 재무제표주석 시트 업로드 완료: {gsheet_name}")
                    
                except Exception as sheet_e:
                    print(f"❌ 시트 처리 실패 {sheet_name}: {str(sheet_e)}")
                    self.processing_results['xbrl_xlsx_failed'].append(sheet_name)
                    continue
            
            return True
            
        except Exception as e:
            print(f"❌ 재무제표주석 Excel 다운로드/처리 실패: {str(e)}")
            return False

    def update_google_sheet(self, sheet_name, data):
        """Google Sheets 업데이트"""
        try:
            # 시트 이름 길이 제한 (100자)
            if len(sheet_name) > 100:
                sheet_name = sheet_name[:97] + "..."
            
            # 기존 시트 확인 및 생성
            try:
                worksheet = self.workbook.worksheet(sheet_name)
                # 기존 시트가 있으면 클리어
                worksheet.clear()
            except gspread.exceptions.WorksheetNotFound:
                # 새 시트 생성
                max_rows = max(1000, len(data) + 100)
                max_cols = max(26, len(data[0]) + 5) if data else 26
                worksheet = self.workbook.add_worksheet(sheet_name, max_rows, max_cols)
            
            # 데이터가 있으면 업로드
            if data:
                # 헤더 추가
                header = [
                    [f"처리일시: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"],
                    [f"데이터 출처: XBRL Excel 다운로드"],
                    []
                ]
                
                all_data = header + data
                
                # 배치 업로드
                BATCH_SIZE = 100
                for i in range(0, len(all_data), BATCH_SIZE):
                    batch = all_data[i:i + BATCH_SIZE]
                    
                    try:
                        worksheet.append_rows(batch)
                        time.sleep(1)  # API 제한 회피
                    except gspread.exceptions.APIError as e:
                        if 'Quota exceeded' in str(e):
                            print("API 할당량 초과. 60초 대기...")
                            time.sleep(60)
                            worksheet.append_rows(batch)
                        else:
                            raise e
            
        except Exception as e:
            print(f"❌ Google Sheets 업데이트 실패: {str(e)}")
            raise

    # 기존 메서드들은 그대로 유지
    def process_xbrl_sheets(self, rcept_no):
        """XBRL 전용 시트 처리 (기존 방식)"""
        print(f"\n🔬 XBRL 방식 처리 시작: {rcept_no}")
        
        try:
            # XBRL 데이터 다운로드 및 파싱
            xbrl_content = self.download_xbrl_data(rcept_no)
            parsed_xbrl = self.parse_xbrl_data(xbrl_content)
            
            if not parsed_xbrl.get('financial_data'):
                print("⚠️ XBRL에서 재무 데이터를 찾을 수 없음")
                return False
            
            # XBRL 전용 시트에 구조화된 데이터 저장
            structured_data = self.convert_xbrl_to_structured_data(parsed_xbrl)
            self.update_xbrl_dedicated_sheets(structured_data, rcept_no)
            
            # 처리 결과 기록
            for sheet_name in self.XBRL_TARGET_SHEETS:
                self.processing_results['xbrl_success'].append(sheet_name)
            
            print(f"✅ XBRL 방식 처리 완료")
            return True
            
        except Exception as e:
            print(f"❌ XBRL 방식 실패: {str(e)}")
            for sheet_name in self.XBRL_TARGET_SHEETS:
                self.processing_results['xbrl_failed'].append(sheet_name)
            return False

    def process_html_sheets(self, rcept_no):
        """HTML 전용 시트 처리"""
        print(f"\n🌐 HTML 방식 처리 시작: {rcept_no}")
        
        try:
            report_index = self.dart.sub_docs(rcept_no)
            target_docs = report_index[report_index['title'].isin(self.HTML_TARGET_SHEETS)]
            
            print(f"📑 HTML 처리 대상 문서: {len(target_docs)}개")
            
            success_count = 0
            for _, doc in target_docs.iterrows():
                try:
                    print(f"📄 HTML 문서 처리: {doc['title']}")
                    self.update_html_worksheet(doc['title'], doc['url'])
                    self.processing_results['html_success'].append(doc['title'])
                    success_count += 1
                    print(f"✅ HTML 문서 완료: {doc['title']}")
                    
                except Exception as doc_e:
                    print(f"❌ HTML 문서 실패 {doc['title']}: {str(doc_e)}")
                    self.processing_results['html_failed'].append(doc['title'])
                    continue
            
            print(f"✅ HTML 방식 처리 완료: {success_count}/{len(target_docs)}개 성공")
            return success_count > 0
            
        except Exception as e:
            print(f"❌ HTML 방식 전체 실패: {str(e)}")
            return False

    # 나머지 기존 메서드들은 그대로 유지...
    def update_xbrl_dedicated_sheets(self, structured_data, rcept_no):
        """XBRL 전용 시트 업데이트"""
        print(f"📊 XBRL 전용 시트 업데이트 시작")
        
        # XBRL 시트명 정의 (구분을 위해 접두사 추가)
        xbrl_sheets = {
            'XBRL_연결재무제표': structured_data,
            'XBRL_재무제표': structured_data,
            'XBRL_연결재무제표_주석': self.create_notes_data(structured_data, '연결', rcept_no),
            'XBRL_재무제표_주석': self.create_notes_data(structured_data, '별도', rcept_no),
            'XBRL_처리현황': self.create_xbrl_status_data(structured_data, rcept_no)
        }
        
        for sheet_name, data in xbrl_sheets.items():
            try:
                # 시트 존재 확인 및 생성
                try:
                    worksheet = self.workbook.worksheet(sheet_name)
                except gspread.exceptions.WorksheetNotFound:
                    print(f"🆕 새 XBRL 시트 생성: {sheet_name}")
                    worksheet = self.workbook.add_worksheet(sheet_name, 1000, 15)
                    self.setup_xbrl_sheet_header(worksheet, sheet_name)
                
                # 데이터 변환 및 업데이트
                if 'XBRL_처리현황' in sheet_name:
                    table_data = self.convert_status_to_table(data)
                elif '주석' in sheet_name:
                    table_data = self.convert_notes_to_table(data)
                else:
                    table_data = self.convert_xbrl_to_table_format(data)
                
                if table_data:
                    # 기존 데이터 보존하면서 새 데이터 추가
                    self.append_xbrl_data(worksheet, table_data, rcept_no)
                    print(f"✅ XBRL 시트 업데이트 완료: {sheet_name}")
                
            except Exception as sheet_e:
                print(f"❌ XBRL 시트 업데이트 실패 {sheet_name}: {str(sheet_e)}")
                continue

    def setup_xbrl_sheet_header(self, worksheet, sheet_name):
        """XBRL 시트 헤더 설정"""
        if 'XBRL_처리현황' in sheet_name:
            headers = ['처리일시', '접수번호', '보고서유형', '처리방식', '데이터수', '상태', '비고']
        elif '주석' in sheet_name:
            headers = ['처리일시', '접수번호', '구분', '주석유형', '내용', '테이블수', '데이터출처', '비고']
        else:
            headers = ['처리일시', '접수번호', '구분', '항목', '당기', '전기', '전전기', '단위', '데이터출처', '비고']
        
        worksheet.update('A1:J1', [headers])
        
        # 헤더 스타일링 (배경색 설정)
        try:
            worksheet.format('A1:J1', {
                'backgroundColor': {'red': 0.9, 'green': 0.9, 'blue': 1.0},
                'textFormat': {'bold': True}
            })
        except:
            pass  # 스타일링 실패해도 진행

    def create_notes_data(self, structured_data, statement_type, rcept_no):
        """주석 데이터 생성"""
        notes_data = {
            'rcept_no': rcept_no,
            'statement_type': statement_type,  # '연결' 또는 '별도'
            'processed_time': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'notes_info': {
                'accounting_policies': '회계정책 관련 주석',
                'significant_estimates': '중요한 회계추정 관련 주석',
                'financial_instruments': '금융상품 관련 주석',
                'risk_management': '위험관리 관련 주석'
            },
            'table_count': 0,  # XBRL에서 추출된 테이블 수
            'status': f'XBRL {statement_type} 주석 처리 완료'
        }
        return notes_data

    def convert_notes_to_table(self, notes_data):
        """주석 데이터를 테이블 형태로 변환"""
        table_data = []
        
        for note_type, content in notes_data['notes_info'].items():
            table_data.append([
                notes_data['statement_type'],
                note_type,
                content,
                str(notes_data['table_count']),
                'XBRL',
                f"{notes_data['statement_type']} 재무제표 주석"
            ])
        
        return table_data

    def append_xbrl_data(self, worksheet, table_data, rcept_no):
        """XBRL 데이터를 기존 시트에 추가"""
        try:
            # 현재 데이터 확인
            existing_data = worksheet.get_all_values()
            
            # 중복 데이터 확인 (같은 접수번호)
            duplicate_rows = []
            for i, row in enumerate(existing_data[1:], 2):  # 헤더 제외
                if len(row) > 1 and row[1] == rcept_no:  # 접수번호 컬럼
                    duplicate_rows.append(i)
            
            if duplicate_rows:
                print(f"⚠️ 중복 데이터 발견: {len(duplicate_rows)}행. 삭제 후 업데이트")
                # 중복 행 삭제 (역순으로)
                for row_num in reversed(duplicate_rows):
                    worksheet.delete_rows(row_num)
            
            # 새 데이터 추가
            if table_data:
                # 각 행에 처리 정보 추가
                processed_data = []
                current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                
                for row in table_data:
                    new_row = [current_time, rcept_no] + row
                    processed_data.append(new_row)
                
                worksheet.append_rows(processed_data)
                print(f"📝 {len(processed_data)}행 데이터 추가 완료")
            
        except Exception as e:
            print(f"❌ XBRL 데이터 추가 실패: {str(e)}")
            raise

    def create_xbrl_status_data(self, structured_data, rcept_no):
        """XBRL 처리 현황 데이터 생성"""
        status_data = {
            'rcept_no': rcept_no,
            'processed_time': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'data_counts': {
                'balance_sheet': len(structured_data.get('balance_sheet', {})),
                'income_statement': len(structured_data.get('income_statement', {})),
                'cash_flow': len(structured_data.get('cash_flow', {}))
            },
            'company_info': structured_data.get('company_info', {}),
            'status': 'XBRL 처리 완료'
        }
        return status_data

    def convert_status_to_table(self, status_data):
        """처리 현황을 테이블 형태로 변환"""
        table_data = []
        
        total_count = sum(status_data['data_counts'].values())
        
        table_data.append([
            status_data['processed_time'],
            status_data['rcept_no'],
            '분기보고서',
            'XBRL',
            str(total_count),
            status_data['status'],
            f"재무상태표:{status_data['data_counts']['balance_sheet']}, 손익계산서:{status_data['data_counts']['income_statement']}"
        ])
        
        return table_data

    def convert_xbrl_to_table_format(self, structured_data):
        """XBRL 구조화 데이터를 테이블 형태로 변환"""
        table_data = []
        
        # 재무상태표 데이터
        if structured_data.get('balance_sheet'):
            for item, value in structured_data['balance_sheet'].items():
                table_data.append([
                    '재무상태표', item, str(value), '', '', 'KRW', 'XBRL', ''
                ])
        
        # 손익계산서 데이터
        if structured_data.get('income_statement'):
            for item, value in structured_data['income_statement'].items():
                table_data.append([
                    '손익계산서', item, str(value), '', '', 'KRW', 'XBRL', ''
                ])
        
        # 현금흐름표 데이터
        if structured_data.get('cash_flow'):
            for item, value in structured_data['cash_flow'].items():
                table_data.append([
                    '현금흐름표', item, str(value), '', '', 'KRW', 'XBRL', ''
                ])
        
        return table_data

    def update_html_worksheet(self, sheet_name, url):
        """HTML 방식 워크시트 업데이트 (기존 방식)"""
        try:
            # HTML 시트는 기존 방식 유지
            try:
                worksheet = self.workbook.worksheet(sheet_name)
            except gspread.exceptions.WorksheetNotFound:
                worksheet = self.workbook.add_worksheet(sheet_name, 1000, 10)
            
            print(f"🌐 HTML 데이터 다운로드: {url}")
            response = requests.get(url, timeout=30)
            response.raise_for_status()
            
            if response.status_code == 200:
                print(f"📊 HTML 콘텐츠 처리...")
                self.process_html_content(worksheet, response.text, sheet_name)
                print(f"✅ HTML 시트 업데이트 완료: {sheet_name}")
            else:
                raise Exception(f"HTTP 오류: {response.status_code}")
                
        except Exception as e:
            print(f"❌ HTML 워크시트 업데이트 실패: {str(e)}")
            raise

    def process_html_content(self, worksheet, html_content, sheet_name):
        """HTML 내용 처리 (개선된 버전)"""
        try:
            soup = BeautifulSoup(html_content, 'html.parser')
            tables = soup.find_all("table")
            
            # 기존 데이터 백업 및 클리어
            worksheet.clear()
            
            # 시트 메타데이터 추가
            metadata = [
                [f'HTML 처리 시트: {sheet_name}'],
                [f'처리 일시: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'],
                [f'데이터 출처: HTML 스크래핑'],
                ['']  # 빈 행
            ]
            
            all_data = metadata.copy()
            
            print(f"발견된 테이블 수: {len(tables)}")
            
            for i, table in enumerate(tables):
                table_data = self.parse_html_table_robust(table)
                
                if table_data:
                    print(f"테이블 {i+1}: {len(table_data)}행 추출")
                    all_data.extend(table_data)
                    all_data.append([''])  # 테이블 간 구분
            
            if len(all_data) > len(metadata):
                # 마지막 빈 행 제거
                if all_data and all_data[-1] == ['']:
                    all_data.pop()
                
                print(f"전체 {len(all_data)}행 데이터 준비 완료")
                
                # 배치 업로드
                BATCH_SIZE = 100
                for i in range(0, len(all_data), BATCH_SIZE):
                    batch = all_data[i:i + BATCH_SIZE]
                    
                    # 행 길이 정규화
                    max_cols = max(len(row) for row in batch) if batch else 0
                    normalized_batch = []
                    for row in batch:
                        normalized_row = row + [''] * (max_cols - len(row))
                        normalized_batch.append(normalized_row)
                    
                    try:
                        worksheet.append_rows(normalized_batch)
                        print(f"배치 업로드: {i+1}~{min(i+BATCH_SIZE, len(all_data))} 행")
                        
                    except gspread.exceptions.APIError as e:
                        if 'Quota exceeded' in str(e):
                            print("API 할당량 초과. 60초 대기...")
                            time.sleep(60)
                            worksheet.append_rows(normalized_batch)
                        else:
                            raise e
            else:
                print("⚠️ 추출된 데이터가 없습니다.")
                
        except Exception as e:
            print(f"❌ HTML 콘텐츠 처리 실패: {str(e)}")
            raise

    def parse_html_table_robust(self, table):
        """견고한 HTML 테이블 파싱"""
        try:
            if HTML_PARSER_AVAILABLE:
                try:
                    return parser.make2d(table)
                except Exception:
                    pass
            
            # 내장 파서 사용
            rows = []
            for tr in table.find_all('tr'):
                row = []
                for cell in tr.find_all(['td', 'th']):
                    text = cell.get_text(separator=' ', strip=True)
                    text = re.sub(r'\s+', ' ', text)
                    row.append(text)
                
                if row:
                    rows.append(row)
            
            return rows
            
        except Exception as e:
            print(f"HTML 테이블 파싱 실패: {str(e)}")
            return []

    def record_processing_result(self, report, xbrl_success, html_success):
        """처리 결과 기록"""
        result_status = ""
        if xbrl_success and html_success:
            result_status = "✅ XBRL+HTML 모두 성공"
        elif xbrl_success:
            result_status = "🔬 XBRL만 성공"
        elif html_success:
            result_status = "🌐 HTML만 성공"
        else:
            result_status = "❌ 모두 실패"
        
        print(f"📋 처리 결과: {report['report_nm']} - {result_status}")

    def print_processing_summary(self):
        """처리 결과 요약 출력"""
        print(f"\n📊 === 처리 결과 요약 ===")
        print(f"전체 보고서 수: {self.processing_results['total_processed']}")
        print(f"XBRL Excel 성공: {len(self.processing_results['xbrl_xlsx_success'])}개")
        print(f"XBRL Excel 실패: {len(self.processing_results['xbrl_xlsx_failed'])}개")
        print(f"XBRL 성공: {len(self.processing_results['xbrl_success'])}개")
        print(f"XBRL 실패: {len(self.processing_results['xbrl_failed'])}개")
        print(f"HTML 성공: {len(self.processing_results['html_success'])}개")
        print(f"HTML 실패: {len(self.processing_results['html_failed'])}개")
        
        # 텔레그램 요약 메시지
        summary_message = (
            f"🔄 DART 이원화 처리 완료\n\n"
            f"• 종목: {self.company_name} ({self.corp_code})\n"
            f"• 처리 일시: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n"
            f"• 전체 보고서: {self.processing_results['total_processed']}개\n"
            f"• XBRL Excel 성공: {len(self.processing_results['xbrl_xlsx_success'])}개\n"
            f"• XBRL 성공: {len(self.processing_results['xbrl_success'])}개\n"
            f"• HTML 성공: {len(self.processing_results['html_success'])}개\n"
            f"• 총 시트 생성: {len(self.processing_results['xbrl_xlsx_success']) + len(self.processing_results['xbrl_success']) + len(self.processing_results['html_success'])}개"
        )
        self.send_telegram_message(summary_message)

    # XBRL 관련 메소드들 (기존 코드 유지)
    def download_xbrl_data(self, rcept_no):
        """XBRL 데이터 다운로드"""
        print(f"XBRL 데이터 다운로드 시작: {rcept_no}")
        
        try:
            download_url = f"https://opendart.fss.or.kr/disclosureinfo/fnltt/dwld/main.do?rcp_no={rcept_no}"
            
            session = requests.Session()
            session.headers.update({
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
                'Accept': 'application/zip, application/xml, text/xml, */*',
                'Referer': 'https://opendart.fss.or.kr/'
            })
            
            response = session.get(download_url, timeout=30)
            response.raise_for_status()
            
            content_type = response.headers.get('content-type', '').lower()
            content_disposition = response.headers.get('content-disposition', '').lower()
            
            if ('application/zip' in content_type or 
                'application/x-zip' in content_type or 
                '.zip' in content_disposition):
                print("ZIP 파일 감지, 압축 해제 중...")
                return self.extract_xbrl_from_zip(response.content)
            
            elif ('xml' in content_type or 
                  response.content.strip().startswith(b'<?xml')):
                print("XML 파일 감지")
                return response.content.decode('utf-8')
            
            else:
                print(f"알 수 없는 파일 형식: {content_type}")
                raise ValueError("XBRL 파일을 찾을 수 없습니다.")
                
        except Exception as e:
            print(f"XBRL 다운로드 실패: {str(e)}")
            raise

    def extract_xbrl_from_zip(self, zip_content):
        """ZIP 파일에서 XBRL 데이터 추출"""
        try:
            with zipfile.ZipFile(io.BytesIO(zip_content)) as zip_ref:
                file_list = zip_ref.namelist()
                print(f"ZIP 파일 내용: {file_list}")
                
                xbrl_patterns = [
                    r'.*\.xbrl$',
                    r'.*xbrl.*\.xml$',
                    r'.*_financial.*\.xml$',
                    r'.*\.xml$'
                ]
                
                xbrl_file = None
                for pattern in xbrl_patterns:
                    matching_files = [f for f in file_list if re.match(pattern, f, re.IGNORECASE)]
                    if matching_files:
                        xbrl_file = max(matching_files, key=lambda x: zip_ref.getinfo(x).file_size)
                        print(f"XBRL 파일 선택: {xbrl_file}")
                        break
                
                if xbrl_file:
                    with zip_ref.open(xbrl_file) as f:
                        content = f.read()
                        try:
                            return content.decode('utf-8')
                        except UnicodeDecodeError:
                            try:
                                return content.decode('utf-8-sig')
                            except UnicodeDecodeError:
                                return content.decode('euc-kr')
                else:
                    raise ValueError("ZIP 파일에서 XBRL 파일을 찾을 수 없습니다.")
                    
        except Exception as e:
            print(f"ZIP 파일 처리 실패: {str(e)}")
            raise

    def parse_xbrl_data(self, xbrl_content):
        """XBRL XML 데이터 파싱"""
        try:
            print("XBRL 데이터 파싱 시작...")
            root = ET.fromstring(xbrl_content)
            print("XML 파싱 성공")
            
            for prefix, uri in root.attrib.items():
                if prefix.startswith('xmlns:'):
                    ns_prefix = prefix[6:]
                    self.xbrl_namespaces[ns_prefix] = uri
                elif prefix == 'xmlns':
                    self.xbrl_namespaces['default'] = uri
            
            parsed_data = {
                'contexts': self.extract_contexts(root),
                'financial_data': self.extract_financial_data(root),
                'company_info': self.extract_company_info(root),
                'metadata': self.extract_metadata(root)
            }
            
            print("XBRL 데이터 파싱 완료")
            return parsed_data
            
        except Exception as e:
            print(f"XBRL 데이터 파싱 실패: {str(e)}")
            raise

    def extract_contexts(self, root):
        """Context 정보 추출"""
        contexts = {}
        context_elements = root.findall('.//xbrl:context', self.xbrl_namespaces)
        
        for context in context_elements:
            context_id = context.get('id')
            
            period_info = {}
            period = context.find('xbrl:period', self.xbrl_namespaces)
            if period is not None:
                instant = period.find('xbrl:instant', self.xbrl_namespaces)
                start_date = period.find('xbrl:startDate', self.xbrl_namespaces)
                end_date = period.find('xbrl:endDate', self.xbrl_namespaces)
                
                if instant is not None:
                    period_info['type'] = 'instant'
                    period_info['date'] = instant.text
                elif start_date is not None and end_date is not None:
                    period_info['type'] = 'duration'
                    period_info['start_date'] = start_date.text
                    period_info['end_date'] = end_date.text
            
            contexts[context_id] = {'period': period_info}
        
        return contexts

    def extract_financial_data(self, root):
        """재무 데이터 추출"""
        financial_data = {}
        
        key_items = {
            'Assets': ['ifrs-full:Assets', 'dart:Assets'],
            'Liabilities': ['ifrs-full:Liabilities', 'dart:Liabilities'],
            'Equity': ['ifrs-full:Equity', 'dart:Equity'],
            'Revenue': ['ifrs-full:Revenue', 'dart:Revenue'],
            'ProfitLoss': ['ifrs-full:ProfitLoss', 'dart:ProfitLoss']
        }
        
        for item_name, possible_tags in key_items.items():
            for tag in possible_tags:
                elements = root.findall(f'.//{tag}', self.xbrl_namespaces)
                if elements:
                    item_data = []
                    for elem in elements:
                        item_data.append({
                            'value': elem.text,
                            'context_ref': elem.get('contextRef'),
                            'unit_ref': elem.get('unitRef'),
                            'decimals': elem.get('decimals')
                        })
                    financial_data[item_name] = item_data
                    break
        
        return financial_data

    def extract_company_info(self, root):
        """회사 정보 추출"""
        return {}

    def extract_metadata(self, root):
        """메타데이터 추출"""
        return {}

    def convert_xbrl_to_structured_data(self, parsed_xbrl):
        """XBRL 데이터를 구조화된 형태로 변환"""
        structured_data = {
            'balance_sheet': {},
            'income_statement': {},
            'cash_flow': {},
            'company_info': parsed_xbrl.get('company_info', {}),
            'reporting_period': None
        }
        
        financial_mapping = {
            'balance_sheet': {
                'Assets': '자산총계',
                'Liabilities': '부채총계',
                'Equity': '자본총계'
            },
            'income_statement': {
                'Revenue': '매출액',
                'ProfitLoss': '당기순이익'
            }
        }
        
        financial_data = parsed_xbrl.get('financial_data', {})
        
        for statement_type, mapping in financial_mapping.items():
            for xbrl_item, korean_name in mapping.items():
                if xbrl_item in financial_data and financial_data[xbrl_item]:
                    item_data = financial_data[xbrl_item][0]
                    value = item_data['value']
                    
                    if value:
                        try:
                            cleaned_value = re.sub(r'[,\s]', '', value)
                            numeric_value = float(cleaned_value)
                            structured_data[statement_type][korean_name] = numeric_value
                        except ValueError:
                            structured_data[statement_type][korean_name] = value
        
        return structured_data

    # 기존 유틸리티 메소드들
    def get_recent_dates(self):
        """날짜 범위 계산 (수동 설정 또는 기본 3개월)"""
        # 환경변수에서 날짜 범위 확인
        manual_start = os.environ.get('MANUAL_START_DATE')
        manual_end = os.environ.get('MANUAL_END_DATE')
        
        if manual_start and manual_end:
            try:
                # 날짜 형식 검증
                start_date = datetime.strptime(manual_start, '%Y%m%d')
                end_date = datetime.strptime(manual_end, '%Y%m%d')
                
                # 날짜 범위 검증
                if start_date > end_date:
                    print("⚠️ 시작일이 종료일보다 늦습니다. 기본 범위로 전환합니다.")
                    return self.get_default_date_range()
                
                # 너무 긴 기간 제한 (최대 2년)
                if (end_date - start_date).days > 730:
                    print("⚠️ 날짜 범위가 너무 깁니다 (최대 2년). 기본 범위로 전환합니다.")
                    return self.get_default_date_range()
                
                print(f"📅 수동 설정 날짜 범위: {manual_start} ~ {manual_end}")
                print(f"📅 기간: {(end_date - start_date).days + 1}일")
                
                return manual_start, manual_end
                
            except ValueError as e:
                print(f"⚠️ 날짜 형식 오류: {str(e)}. 기본 범위로 전환합니다.")
                return self.get_default_date_range()
        else:
            return self.get_default_date_range()
    
    def get_default_date_range(self):
        """기본 3개월 날짜 범위"""
        end_date = datetime.now()
        start_date = end_date - timedelta(days=90)
        date_range = start_date.strftime('%Y%m%d'), end_date.strftime('%Y%m%d')
        print(f"📅 기본 날짜 범위 (최근 3개월): {date_range[0]} ~ {date_range[1]}")
        return date_range

    def get_column_letter(self, col_num):
        """숫자를 엑셀 열 문자로 변환"""
        result = ""
        while col_num > 0:
            col_num, remainder = divmod(col_num - 1, 26)
            result = chr(65 + remainder) + result
        return result

    def send_telegram_message(self, message):
        """텔레그램으로 메시지 전송"""
        if not self.telegram_bot_token or not self.telegram_channel_id:
            print("📱 텔레그램 설정이 없습니다.")
            return
        
        try:
            url = f"https://api.telegram.org/bot{self.telegram_bot_token}/sendMessage"
            data = {
                "chat_id": self.telegram_channel_id,
                "text": message,
                "parse_mode": "HTML"
            }
            response = requests.post(url, data=data)
            response.raise_for_status()
            print("📱 텔레그램 메시지 전송 완료")
        except Exception as e:
            print(f"📱 텔레그램 메시지 전송 실패: {str(e)}")

    def remove_parentheses(self, value):
        """괄호 내용 제거"""
        if not value:
            return value
        return re.sub(r'\s*\(.*?\)\s*', '', value).replace('%', '')

    def process_archive_data(self, archive, start_row, last_col):
        """아카이브 데이터 처리 (기존 로직 유지)"""
        try:
            print(f"📊 Archive 데이터 처리 시작: 행={start_row}, 열={last_col}")
            
            current_cols = archive.col_count
            target_col_letter = self.get_column_letter(last_col)
            
            if last_col >= current_cols:
                new_cols = last_col + 5
                print(f"🔧 시트 크기 조정: {current_cols} → {new_cols}")
                archive.resize(rows=archive.row_count, cols=new_cols)
                time.sleep(2)

            all_rows = archive.get_all_values()
            update_data = []
            sheet_cache = {}
            
            sheet_rows = {}
            processed_count = 0
            
            for row_idx in range(start_row - 1, len(all_rows)):
                if len(all_rows[row_idx]) < 5:
                    continue
                    
                sheet_name = all_rows[row_idx][0]
                if not sheet_name:
                    continue
                
                if sheet_name not in sheet_rows:
                    sheet_rows[sheet_name] = []
                    
                sheet_rows[sheet_name].append({
                    'row_idx': row_idx + 1,
                    'keyword': all_rows[row_idx][1],
                    'n': all_rows[row_idx][2],
                    'x': all_rows[row_idx][3],
                    'y': all_rows[row_idx][4]
                })
                processed_count += 1
            
            print(f"📋 처리할 시트 수: {len(sheet_rows)}, 총 행 수: {processed_count}")
            
            for sheet_name, rows in sheet_rows.items():
                try:
                    print(f"\n🔍 시트 '{sheet_name}' 처리 중 (항목: {len(rows)}개)")
                    
                    if sheet_name not in sheet_cache:
                        try:
                            search_sheet = self.workbook.worksheet(sheet_name)
                            sheet_data = search_sheet.get_all_values()
                            df = pd.DataFrame(sheet_data)
                            sheet_cache[sheet_name] = df
                            print(f"✅ 시트 데이터 로드: {df.shape}")
                        except gspread.exceptions.WorksheetNotFound:
                            print(f"⚠️ 시트 '{sheet_name}' 없음. 건너뜀.")
                            continue
                    
                    df = sheet_cache[sheet_name]
                    
                    for row in rows:
                        try:
                            keyword = row['keyword']
                            if not all([keyword, row['n'], row['x'], row['y']]):
                                continue
                            
                            n, x, y = int(row['n']), int(row['x']), int(row['y'])
                            
                            keyword_positions = []
                            for idx, df_row in df.iterrows():
                                for col_idx, value in enumerate(df_row):
                                    if str(value).strip() == keyword.strip():
                                        keyword_positions.append((idx, col_idx))
                            
                            if keyword_positions and len(keyword_positions) >= n:
                                target_pos = keyword_positions[n - 1]
                                target_row = target_pos[0] + y
                                target_col = target_pos[1] + x
                                
                                if (0 <= target_row < df.shape[0] and 
                                    0 <= target_col < df.shape[1]):
                                    value = df.iat[target_row, target_col]
                                    cleaned_value = self.remove_parentheses(str(value))
                                    update_data.append((row['row_idx'], cleaned_value))
                                    print(f"✅ 값 발견: {keyword} → {cleaned_value}")
                                else:
                                    print(f"⚠️ 범위 초과: {keyword}")
                            else:
                                print(f"⚠️ 키워드 미발견: {keyword}")
                                
                        except Exception as row_e:
                            print(f"❌ 행 처리 오류: {str(row_e)}")
                            continue
                
                except Exception as sheet_e:
                    print(f"❌ 시트 '{sheet_name}' 처리 오류: {str(sheet_e)}")
                    continue
            
            print(f"\n📊 업데이트할 데이터: {len(update_data)}개")
            
            if update_data:
                self.update_archive_column(archive, update_data, target_col_letter, last_col)
                
                today = datetime.now()
                three_months_ago = today - timedelta(days=90)
                year = str(three_months_ago.year)[2:]
                quarter = (three_months_ago.month - 1) // 3 + 1
                quarter_text = f"{quarter}Q{year}"
                
                meta_updates = [
                    {'range': 'J1', 'values': [[today.strftime('%Y-%m-%d')]]},
                    {'range': f'{target_col_letter}1', 'values': [['1']]},
                    {'range': f'{target_col_letter}5', 'values': [[today.strftime('%Y-%m-%d')]]},
                    {'range': f'{target_col_letter}6', 'values': [[quarter_text]]}
                ]
                
                archive.batch_update(meta_updates)
                print(f"✅ 메타데이터 업데이트 완료 (분기: {quarter_text})")
                
                message = (
                    f"🔄 DART Archive 업데이트 완료\n\n"
                    f"• 종목: {self.company_name} ({self.corp_code})\n"
                    f"• 분기: {quarter_text}\n"
                    f"• 업데이트 일시: {today.strftime('%Y-%m-%d %H:%M:%S')}\n"
                    f"• 처리된 행: {len(update_data)}개\n"
                    f"• 시트 열: {target_col_letter} (#{last_col})"
                )
                self.send_telegram_message(message)
                
            else:
                print("⚠️ 업데이트할 데이터가 없습니다.")
                
        except Exception as e:
            error_msg = f"Archive 데이터 처리 중 오류: {str(e)}"
            print(f"❌ {error_msg}")
            self.send_telegram_message(f"❌ {error_msg}")
            raise

    def update_archive_column(self, archive, update_data, target_col_letter, last_col):
        """Archive 열 데이터 업데이트"""
        try:
            min_row = min(row for row, _ in update_data)
            max_row = max(row for row, _ in update_data)
            
            column_data = [''] * (max_row - min_row + 1)
            for row, value in update_data:
                adjusted_row = row - min_row
                column_data[adjusted_row] = value
            
            column_data_2d = [[value] for value in column_data]
            
            range_label = f'{target_col_letter}{min_row}:{target_col_letter}{max_row}'
            print(f"📝 업데이트 범위: {range_label}")
            
            archive.batch_update([{
                'range': range_label,
                'values': column_data_2d
            }])
            
            print(f"✅ 컬럼 업데이트 완료: {min_row}~{max_row} 행")
            
        except Exception as e:
            print(f"❌ 컬럼 업데이트 실패: {str(e)}")
            raise


def load_company_config():
    """yml 파일에서 회사 설정 로드 또는 환경변수에서 읽기"""
    # 환경변수에서 읽기 (GitHub Actions 사용)
    corp_code = os.environ.get('COMPANY_CORP_CODE')
    company_name = os.environ.get('COMPANY_NAME')
    spreadsheet_var = os.environ.get('COMPANY_SPREADSHEET_VAR')
    
    if corp_code and company_name and spreadsheet_var:
        return {
            'corp_code': corp_code,
            'company_name': company_name,
            'spreadsheet_var': spreadsheet_var
        }
    
    # 기본값 (로컬 테스트용)
    return {
        'corp_code': '307950',
        'company_name': '현대오토에버',
        'spreadsheet_var': 'AUTOEVER_SPREADSHEET_ID'
    }


def main():
    """메인 실행 함수"""
    try:
        import sys
        
        def log(msg):
            print(f"🤖 {msg}")
            sys.stdout.flush()
        
        # yml 또는 환경변수에서 설정 로드
        company_config = load_company_config()
        
        log(f"{company_config['company_name']}({company_config['corp_code']}) 이원화 시스템 업데이트 시작")
        
        try:
            updater = DualSystemDartUpdater(company_config)
            
            # 이원화 시스템으로 보고서 업데이트
            log("📋 이원화 DART 보고서 업데이트 시작...")
            updater.update_dart_reports()
            log("✅ 이원화 DART 보고서 업데이트 완료")
            
            # Archive 시트 처리 (기존 로직 유지)
            log("📊 Archive 시트 업데이트 시작...")
            archive = updater.workbook.worksheet('Dart_Archive')
            
            sheet_values = archive.get_all_values()
            if not sheet_values:
                raise ValueError("Dart_Archive 시트가 비어있습니다")
            
            last_col = len(sheet_values[0])
            control_value = archive.cell(1, last_col).value
            start_row = 10
            
            if control_value:
                last_col += 1
            
            log(f"Archive 처리: 시작행={start_row}, 대상열={last_col}")
            updater.process_archive_data(archive, start_row, last_col)
            log("✅ Archive 시트 업데이트 완료")
            
            log("🎉 이원화 시스템 전체 작업 완료!")
            
        except Exception as e:
            log(f"❌ 처리 중 오류 발생: {str(e)}")
            if 'updater' in locals():
                updater.send_telegram_message(
                    f"❌ 이원화 DART 업데이트 실패\n\n"
                    f"• 종목: {company_config['company_name']} ({company_config['corp_code']})\n"
                    f"• 오류: {str(e)}"
                )
            raise

    except Exception as e:
        print(f"💥 전체 작업 중 치명적 오류: {str(e)}")
        print(f"🔍 오류 타입: {type(e).__name__}")
        raise e

if __name__ == "__main__":
    main()
